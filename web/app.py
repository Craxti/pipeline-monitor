"""
FastAPI web interface for CI/CD Monitor.

Run with:  uvicorn web.app:app --host 0.0.0.0 --port 8080 --reload
Or via:    python ci_monitor.py web
"""

from __future__ import annotations

import asyncio
import csv
import html
import io
import json
import logging
import os
import uuid
import shutil
import subprocess
import sys
import threading
import time
from collections import Counter, defaultdict, deque
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote, urljoin, urlparse

import httpx
import yaml
from fastapi import FastAPI, Request, HTTPException, Depends, Header
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from starlette.staticfiles import StaticFiles


from models.models import (
    CISnapshot,
    TestRecord,
    normalize_build_status,
    normalize_service_status,
    normalize_test_status,
)

from config_migrations import migrate_telegram_notifications
from web.schemas import MonitorGeneralConfig

logger = logging.getLogger(__name__)

# ── Shared token auth (optional) ────────────────────────────────────────────

def _shared_api_token(cfg: Optional[dict] = None) -> str:
    """
    Shared token for protecting sensitive endpoints.
    Sources (highest priority first):
    - env: CICD_MON_API_TOKEN
    - config.yaml: web.api_token
    If empty, auth is considered disabled (backwards-compatible).
    """
    env_tok = (os.getenv("CICD_MON_API_TOKEN") or "").strip()
    if env_tok:
        return env_tok
    if not cfg:
        try:
            cfg = _load_yaml_config()
        except Exception:
            cfg = None
    if cfg:
        w = cfg.get("web", {}) or {}
        tok = (w.get("api_token") or "").strip()
        if tok:
            return tok
    return ""


def _token_from_headers(x_api_token: Optional[str], authorization: Optional[str]) -> str:
    if x_api_token:
        return str(x_api_token).strip()
    if authorization:
        raw = str(authorization).strip()
        if raw.lower().startswith("bearer "):
            return raw.split(" ", 1)[1].strip()
    return ""


async def require_shared_token(
    request: Request,
    x_api_token: Optional[str] = Header(default=None, alias="X-API-Token"),
    authorization: Optional[str] = Header(default=None, alias="Authorization"),
) -> None:
    cfg = None
    try:
        cfg = _load_yaml_config()
    except Exception:
        cfg = None
    expected = _shared_api_token(cfg)
    if not expected:
        # Backwards-compatible: if token is not configured, do not block.
        return
    provided = _token_from_headers(x_api_token, authorization)
    if not provided or provided != expected:
        raise HTTPException(status_code=401, detail="Unauthorized")

# Bumped when API surface changes; visible in /api/chat/status so you know the process reloaded.
_APP_BUILD = "2026-04-03+multi-telegram-ollama"

# Shown when provider=cursor but no agent binary/bundle was resolved (before calling the proxy).
CURSOR_AGENT_UNAVAILABLE_MSG = (
    "Cursor Agent не найден на этом компьютере (ни в PATH, ни в настройке «Путь к Cursor Agent», "
    "ни после авто-поиска по типичным папкам). Без отдельного пакета Cursor Agent CLI чат через "
    "cursor-api-proxy работать не будет — редактор Cursor его не подставляет. "
    "Варианты: установить CLI по документации https://cursor.com/docs/cli/overview , "
    "либо указать каталог с agent.cmd + node.exe + index.js в Настройках → AI, "
    "либо переключить провайдера на Gemini или OpenRouter. Лог: data/cursor_proxy.log"
)

_cursor_agent_resolve_cache: tuple[float, str | None] | None = None

# SQLite layer (optional, initialized in lifespan)
try:
    from web.db import (
        init_db,
        append_snapshot as _db_append,
        db_stats,
        service_uptime as _db_svc_uptime,
        build_duration_history as _db_build_duration,
        flaky_analysis as _db_flaky_analysis,
        query_builds_history as _db_query_builds_history,
    )
    _SQLITE_AVAILABLE = True
except ImportError:
    try:
        from db import (
            init_db,
            append_snapshot as _db_append,
            db_stats,
            service_uptime as _db_svc_uptime,
            build_duration_history as _db_build_duration,
            flaky_analysis as _db_flaky_analysis,
            query_builds_history as _db_query_builds_history,
        )
        _SQLITE_AVAILABLE = True
    except ImportError:
        _SQLITE_AVAILABLE = False
        _db_build_duration = None  # type: ignore
        _db_flaky_analysis = None  # type: ignore
        _db_query_builds_history = None  # type: ignore
        logger.debug("SQLite module (db.py) not found — running without persistent history")

# ── data helpers ──────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parent.parent
_TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))
_DATA_FILE = Path("data") / "snapshot.json"
_EVENT_FEED_FILE = Path("data") / "event_feed.json"
_EVENT_FEED_MAX = 500

# In-memory snapshot cache: short TTL + invalidation on _data_revision / file mtime (less disk I/O on polling).
_SNAPSHOT_CACHE_TTL_SEC = 2.0
_snapshot_cache_snap: CISnapshot | None = None
_snapshot_cache_rev: int = -1
_snapshot_cache_mtime: float | None = None
_snapshot_cache_expires_mono: float = 0.0


def _invalidate_snapshot_cache() -> None:
    global _snapshot_cache_snap, _snapshot_cache_rev, _snapshot_cache_mtime, _snapshot_cache_expires_mono
    _snapshot_cache_snap = None
    _snapshot_cache_rev = -1
    _snapshot_cache_mtime = None
    _snapshot_cache_expires_mono = 0.0


def _prime_snapshot_cache(snapshot: CISnapshot, mtime: float | None = None) -> None:
    global _snapshot_cache_snap, _snapshot_cache_rev, _snapshot_cache_mtime, _snapshot_cache_expires_mono
    _snapshot_cache_snap = snapshot
    _snapshot_cache_rev = _data_revision
    if mtime is None:
        try:
            _snapshot_cache_mtime = _DATA_FILE.stat().st_mtime
        except OSError:
            _snapshot_cache_mtime = None
    else:
        _snapshot_cache_mtime = mtime
    _snapshot_cache_expires_mono = time.monotonic() + _SNAPSHOT_CACHE_TTL_SEC


def _config_yaml_path() -> Path:
    """Resolve config.yaml: prefer repo root (next to web/), else CWD (uvicorn odd cwd)."""
    root_cfg = _REPO_ROOT / "config.yaml"
    if root_cfg.is_file():
        return root_cfg
    cwd_cfg = Path("config.yaml")
    if cwd_cfg.is_file():
        return cwd_cfg.resolve()
    return root_cfg


def _load_snapshot() -> CISnapshot | None:
    if not _DATA_FILE.exists():
        _invalidate_snapshot_cache()
        return None
    try:
        st = _DATA_FILE.stat()
    except OSError:
        _invalidate_snapshot_cache()
        return None
    mtime = st.st_mtime
    mon = time.monotonic()
    if (
        _snapshot_cache_snap is not None
        and _snapshot_cache_rev == _data_revision
        and _snapshot_cache_mtime == mtime
        and mon < _snapshot_cache_expires_mono
    ):
        return _snapshot_cache_snap
    try:
        snap = CISnapshot.model_validate_json(_DATA_FILE.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.error("Failed to load snapshot: %s", exc)
        _invalidate_snapshot_cache()
        return None
    _prime_snapshot_cache(snap, mtime)
    return snap


async def _load_snapshot_async() -> CISnapshot | None:
    # Snapshot read+parse can be expensive; keep event loop responsive.
    return await asyncio.to_thread(_load_snapshot)


_HISTORY_FILE = Path("data") / "trends.json"
_HISTORY_MAX_DAYS = 30


def _append_trends(snapshot: CISnapshot) -> None:
    """Append a daily summary bucket to trends.json (one entry per day)."""
    now = datetime.now(tz=timezone.utc)
    day_key = now.strftime("%Y-%m-%d")

    try:
        history: list[dict] = json.loads(_HISTORY_FILE.read_text(encoding="utf-8")) if _HISTORY_FILE.exists() else []
    except Exception:
        history = []

    # Remove today's existing entry (will be replaced with fresh data)
    history = [e for e in history if e.get("date") != day_key]

    # Build per-job failure map
    job_failures: dict[str, int] = {}
    job_totals: dict[str, int] = {}
    for b in snapshot.builds:
        j = b.job_name
        job_totals[j] = job_totals.get(j, 0) + 1
        if b.status_normalized in ("failure", "unstable"):
            job_failures[j] = job_failures.get(j, 0) + 1

    # Per-test failure counts
    test_failures: dict[str, int] = {}
    for t in snapshot.tests:
        if t.status_normalized in ("failed", "error"):
            test_failures[t.test_name] = test_failures.get(t.test_name, 0) + 1

    # Per-source breakdowns (for UI filters in Trends)
    builds_by_source: dict[str, dict[str, int]] = {}
    builds_by_instance: dict[str, dict[str, int]] = {}
    cfg_for_inst = None
    try:
        cfg_for_inst = _load_yaml_config()
    except Exception:
        cfg_for_inst = None

    def _inst_label_for_build_trends(b: object) -> str | None:
        if not cfg_for_inst:
            return None
        return _inst_label_for_build_with_cfg(b, cfg_for_inst)

    for b in snapshot.builds:
        src = str(getattr(b, "source", "") or "").strip().lower() or "unknown"
        st = str(getattr(b, "status_normalized", "") or "").strip().lower()
        rec = builds_by_source.setdefault(src, {"total": 0, "failed": 0})
        rec["total"] += 1
        if st in ("failure", "unstable"):
            rec["failed"] += 1

        inst = _inst_label_for_build_trends(b)
        if inst:
            k = f"{src}|{inst}"
            ir = builds_by_instance.setdefault(k, {"total": 0, "failed": 0})
            ir["total"] += 1
            if st in ("failure", "unstable"):
                ir["failed"] += 1

    def _test_source_group(raw: object) -> str:
        s = str(raw or "").strip().lower()
        if not s:
            return "unknown"
        # Most test data comes from Jenkins console/allure parsers in this project.
        if s.startswith("jenkins"):
            return "jenkins"
        if s.startswith("gitlab"):
            return "gitlab"
        return s

    tests_by_source: dict[str, dict[str, int]] = {}
    test_failures_by_source: dict[str, dict[str, int]] = {}
    for t in snapshot.tests:
        src = _test_source_group(getattr(t, "source", None))
        st = str(getattr(t, "status_normalized", "") or "").strip().lower()
        rec = tests_by_source.setdefault(src, {"total": 0, "failed": 0})
        rec["total"] += 1
        if st in ("failed", "error"):
            rec["failed"] += 1
            tf = test_failures_by_source.setdefault(src, {})
            name = str(getattr(t, "test_name", "") or "")
            tf[name] = tf.get(name, 0) + 1

    # Per-service status snapshot (for uptime history)
    service_health = {s.name: s.status for s in snapshot.services}

    services_down_by_kind: dict[str, int] = {"docker": 0, "http": 0, "other": 0}
    for s in snapshot.services:
        if normalize_service_status(s.status) != "down":
            continue
        kind = str(getattr(s, "kind", "") or "").strip().lower()
        if kind == "docker":
            services_down_by_kind["docker"] += 1
        elif kind == "http":
            services_down_by_kind["http"] += 1
        else:
            services_down_by_kind["other"] += 1

    history.append({
        "date": day_key,
        "ts": now.isoformat(),
        "builds_total": len(snapshot.builds),
        "builds_failed": sum(1 for b in snapshot.builds if b.status_normalized in ("failure", "unstable")),
        "tests_total": len(snapshot.tests),
        "tests_failed": sum(1 for t in snapshot.tests if t.status_normalized in ("failed", "error")),
        "services_down": sum(1 for s in snapshot.services if s.status_normalized == "down"),
        "services_down_by_kind": services_down_by_kind,
        "service_health": service_health,
        "builds_by_source": builds_by_source,
        "builds_by_instance": builds_by_instance,
        "tests_by_source": tests_by_source,
        "job_failures": job_failures,
        "job_totals": job_totals,
        "top_test_failures": sorted(test_failures.items(), key=lambda x: -x[1])[:20],
        "top_test_failures_by_source": {
            src: sorted(m.items(), key=lambda x: -x[1])[:20]
            for src, m in test_failures_by_source.items()
        },
    })

    # Keep only last N days
    cutoff = (now - timedelta(days=_HISTORY_MAX_DAYS)).strftime("%Y-%m-%d")
    history = [e for e in history if e.get("date", "") >= cutoff]
    history.sort(key=lambda e: e.get("date", ""))

    _HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    _HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def save_snapshot(snapshot: CISnapshot) -> None:
    global _data_revision
    with _snapshot_write_lock:
        _DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
        _DATA_FILE.write_text(
            snapshot.model_dump_json(indent=2), encoding="utf-8"
        )
        _data_revision += 1
        try:
            mtime = _DATA_FILE.stat().st_mtime
        except OSError:
            mtime = None
        _prime_snapshot_cache(snapshot, mtime)
    try:
        _append_trends(snapshot)
    except Exception as exc:
        logger.warning("Failed to append trends: %s", exc)
    try:
        _detect_state_changes(snapshot)
    except Exception as exc:
        logger.warning("Failed to detect state changes: %s", exc)
    # Write to SQLite for historical queries (non-blocking, best-effort)
    if _SQLITE_AVAILABLE:
        try:
            _db_append(snapshot)
        except Exception as exc:
            logger.debug("SQLite append skipped: %s", exc)


_snapshot_write_lock = threading.Lock()
_partial_last_write_ts: float = 0.0


def save_snapshot_partial(snapshot: CISnapshot) -> None:
    """
    Persist an in-progress snapshot for live dashboard updates during Collect.
    Intentionally skips trends/notifications/DB to keep it cheap.
    """
    global _data_revision
    # Do not publish a snapshot that wipes tests on disk while a collect is still running
    # and the previous file still had rows — the UI would flash empty (incremental collect
    # clears tests in memory before parsers repopulate).
    try:
        if _collect_state.get("is_collecting"):
            n_new = len(getattr(snapshot, "tests", None) or [])
            if n_new == 0:
                prev = _load_snapshot()
                if prev is not None and len(prev.tests or []) > 0:
                    return
    except Exception:
        pass
    with _snapshot_write_lock:
        _DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
        _DATA_FILE.write_text(snapshot.model_dump_json(indent=2), encoding="utf-8")
        _data_revision += 1
        try:
            mtime = _DATA_FILE.stat().st_mtime
        except OSError:
            mtime = None
        _prime_snapshot_cache(snapshot, mtime)


def _maybe_save_partial(snapshot: CISnapshot, *, min_interval_s: float = 2.0, force: bool = False) -> None:
    global _partial_last_write_ts
    now = time.monotonic()
    if not force and (now - _partial_last_write_ts) < float(min_interval_s):
        return
    _partial_last_write_ts = now
    try:
        save_snapshot_partial(snapshot)
    except Exception as exc:
        logger.debug("Partial snapshot save skipped: %s", exc)


def _normalize_config(cfg: dict) -> dict:
    """Migrate legacy single jenkins/gitlab keys to multi-instance lists."""
    if "jenkins" in cfg and "jenkins_instances" not in cfg:
        inst = dict(cfg.pop("jenkins"))
        inst.setdefault("name", "Jenkins")
        cfg["jenkins_instances"] = [inst]
    if "gitlab" in cfg and "gitlab_instances" not in cfg:
        inst = dict(cfg.pop("gitlab"))
        inst.setdefault("name", "GitLab")
        cfg["gitlab_instances"] = [inst]
    migrate_telegram_notifications(cfg)
    return cfg


def _load_yaml_config() -> dict:
    p = _config_yaml_path()
    if p.is_file():
        with p.open(encoding="utf-8") as fh:
            return _normalize_config(yaml.safe_load(fh) or {})
    return {}


_SETTINGS_SECRET_MASK = "••••••••"


def _is_secret_settings_key(key: str) -> bool:
    lk = key.lower()
    if lk in ("token", "password", "api_key", "bot_token", "private_token", "secret"):
        return True
    if lk == "username":
        return False
    for frag in ("password", "api_key", "secret", "token"):
        if frag in lk:
            return True
    return False


def _mask_settings_for_response(obj: Any) -> Any:
    if isinstance(obj, dict):
        out: dict[str, Any] = {}
        for k, v in obj.items():
            if _is_secret_settings_key(k) and isinstance(v, str) and v.strip():
                out[k] = _SETTINGS_SECRET_MASK
            else:
                out[k] = _mask_settings_for_response(v)
        return out
    if isinstance(obj, list):
        return [_mask_settings_for_response(x) for x in obj]
    return obj


def _merge_settings_secrets(incoming: Any, saved: Any) -> Any:
    """Keep previous secret values when the client sends the mask placeholder or empty."""
    if isinstance(incoming, dict) and isinstance(saved, dict):
        out: dict[str, Any] = {}
        for k, v in incoming.items():
            sv = saved.get(k)
            if _is_secret_settings_key(k) and isinstance(v, str):
                if v == _SETTINGS_SECRET_MASK or (not v.strip() and isinstance(sv, str) and sv.strip()):
                    out[k] = sv if isinstance(sv, str) else v
                else:
                    out[k] = v
            elif isinstance(v, dict) and isinstance(sv, dict):
                out[k] = _merge_settings_secrets(v, sv)
            elif isinstance(v, list) and isinstance(sv, list):
                merged: list[Any] = []
                for i, item in enumerate(v):
                    s_item = sv[i] if i < len(sv) else None
                    if isinstance(item, dict) and isinstance(s_item, dict):
                        merged.append(_merge_settings_secrets(item, s_item))
                    else:
                        merged.append(item)
                out[k] = merged
            else:
                out[k] = v
        return out
    return incoming


def _public_settings_payload(cfg: dict) -> dict:
    """Minimal non-secret fields for UI bootstrapping."""
    g = cfg.get("general") or {}
    w = cfg.get("web") or {}
    sqlite_ok = False
    if _SQLITE_AVAILABLE:
        try:
            st = db_stats()
            sqlite_ok = bool(st.get("enabled"))
        except Exception:
            sqlite_ok = False
    return {
        "ui_language": g.get("ui_language", "en"),
        "project_name": g.get("project_name", "CI/CD Monitor"),
        "web": {
            "host": w.get("host", "0.0.0.0"),
            "port": int(w.get("port", 8000)),
            "auto_collect": w.get("auto_collect", True),
            "collect_interval_seconds": int(w.get("collect_interval_seconds", 300)),
            "live_reload": w.get("live_reload", True),
        },
        "sqlite_enabled": sqlite_ok,
    }


# ── background collection state ───────────────────────────────────────────

_collect_state: dict = {
    "is_collecting": False,
    "last_collected_at": None,   # ISO string or None
    "last_error": None,          # string or None
    "interval_seconds": 300,
    # Live progress (best-effort; updated during collect)
    "started_at": None,          # ISO string or None
    "phase": None,               # jenkins_builds | jenkins_console | jenkins_allure | gitlab | docker | done
    "progress_main": None,       # short line for UI
    "progress_sub": None,        # optional extra line
    "progress_counts": {},       # {"builds": int, "tests": int, "services": int}
}

# Ring buffer of recent collect progress lines for the "Logs" dashboard tab.
_collect_logs: deque[dict[str, Any]] = deque(maxlen=2500)
_collect_slow: deque[dict[str, Any]] = deque(maxlen=800)


def _push_collect_log(phase: str, main: str, sub: str | None = None, level: str = "info") -> None:
    try:
        lvl = (level or "info").strip().lower()
        if lvl not in ("info", "warn", "error"):
            lvl = "info"
        instance: str | None = None
        job: str | None = None
        m = (main or "").strip()
        if m.startswith("Jenkins: "):
            instance = m[len("Jenkins: "):].strip()
        elif m.startswith("GitLab: "):
            instance = m[len("GitLab: "):].strip()
        s = (sub or "").strip()
        if s.startswith("Console: "):
            rest = s[len("Console: "):]
            if " #" in rest:
                job = rest.split(" #", 1)[0].strip()
            else:
                job = rest.strip()
        elif s.startswith("Allure: "):
            rest = s[len("Allure: "):]
            if " #" in rest:
                job = rest.split(" #", 1)[0].strip()
            else:
                job = rest.strip()
        elif s.startswith("Builds: "):
            # Format: "Builds: i/total job_name"
            parts = s.split(" ", 2)
            if len(parts) == 3:
                job = parts[2].strip()
        _collect_logs.append({
            "ts": datetime.now(tz=timezone.utc).isoformat(),
            "level": lvl,
            "phase": phase,
            "main": main,
            "sub": sub,
            "instance": instance,
            "job": job,
            "counts": dict(_collect_state.get("progress_counts") or {}),
        })
    except Exception:
        pass
# Last collect: per-source health (Jenkins / GitLab / Docker monitor)
_instance_health: list[dict[str, Any]] = []
_collect_task: asyncio.Task | None = None
# Server-side switch for the background auto-collect loop.
# We tie it to the UI LIVE mode so when LIVE is off there is no auto-collect.
_auto_collect_enabled: bool = False
_auto_collect_enabled_at_iso: str | None = None

# Bumped on each successful save_snapshot — ETag / cache invalidation / SSE clients
_data_revision: int = 0
_main_loop: asyncio.AbstractEventLoop | None = None
_sse_queues: set[asyncio.Queue] = set()
_MEM_CACHE: dict[str, tuple[float, Any]] = {}
_MEM_CACHE_TTL_SEC = 20.0


def _mem_cache_get(key: str) -> Any | None:
    ent = _MEM_CACHE.get(key)
    if not ent:
        return None
    exp, val = ent
    if time.monotonic() > exp:
        del _MEM_CACHE[key]
        return None
    return val


def _mem_cache_set(key: str, val: Any, ttl: float = _MEM_CACHE_TTL_SEC) -> None:
    _MEM_CACHE[key] = (time.monotonic() + ttl, val)


async def _sse_broadcast_async(payload: dict) -> None:
    for q in list(_sse_queues):
        try:
            q.put_nowait(payload)
        except asyncio.QueueFull:
            try:
                q.get_nowait()
            except Exception:
                pass
            try:
                q.put_nowait(payload)
            except Exception:
                pass


def _status_str(b: object) -> str:
    if isinstance(b, str):
        return b
    return getattr(b, "value", str(b))


def _job_build_analytics(snapshot: CISnapshot) -> dict[str, dict]:
    """Per job: consecutive failures from latest run, last successful build number."""
    from collections import defaultdict

    by_job: dict[str, list] = defaultdict(list)
    for b in snapshot.builds:
        by_job[b.job_name].append(b)

    out: dict[str, dict] = {}
    for job, builds in by_job.items():

        def sort_key(bn: object) -> tuple[float, int]:
            sa = bn.started_at
            if sa is None:
                return (0.0, bn.build_number or 0)
            if sa.tzinfo is None:
                sa = sa.replace(tzinfo=timezone.utc)
            return (sa.timestamp(), bn.build_number or 0)

        builds_sorted = sorted(builds, key=sort_key, reverse=True)
        streak = 0
        for b in builds_sorted:
            if b.status_normalized in ("failure", "unstable"):
                streak += 1
            else:
                break
        last_success_number = None
        for b in builds_sorted:
            if b.status_normalized == "success":
                last_success_number = b.build_number
                break
        latest = builds_sorted[0] if builds_sorted else None
        out[job] = {
            "consecutive_failures": streak,
            "last_success_build_number": last_success_number,
            "latest_status": _status_str(latest.status) if latest else None,
        }
    return out


def _correlation_last_hour() -> dict:
    """Build counts + service state-change events in the last hour (from event_feed)."""
    now = datetime.now(tz=timezone.utc)
    cutoff = now - timedelta(hours=1)
    n_builds = 0
    snap = _load_snapshot()
    if snap:
        for b in snap.builds:
            if not b.started_at:
                continue
            st = b.started_at
            if st.tzinfo is None:
                st = st.replace(tzinfo=timezone.utc)
            else:
                st = st.astimezone(timezone.utc)
            if st >= cutoff:
                n_builds += 1
    n_svc_events = 0
    for e in _event_feed_load(500):
        ts_raw = e.get("ts")
        if not ts_raw:
            continue
        try:
            et = datetime.fromisoformat(str(ts_raw).replace("Z", "+00:00"))
            if et.tzinfo is None:
                et = et.replace(tzinfo=timezone.utc)
            else:
                et = et.astimezone(timezone.utc)
        except Exception:
            continue
        if et < cutoff:
            continue
        k = str(e.get("kind") or "")
        if k.startswith("svc_"):
            n_svc_events += 1
    return {
        "pipelines_started_last_hour": n_builds,
        "service_events_last_hour": n_svc_events,
    }


def _trends_compute(days: int) -> list:
    if not _HISTORY_FILE.exists():
        return []
    try:
        history = json.loads(_HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.error("Failed to load trends: %s", exc)
        return []
    cutoff = (datetime.now(tz=timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    return [e for e in history if e.get("date", "") >= cutoff]


def _uptime_compute(days: int) -> dict:
    if _SQLITE_AVAILABLE:
        try:
            result = _db_svc_uptime(days)
            if result:
                return result
        except Exception:
            pass
    if not _HISTORY_FILE.exists():
        return {}
    try:
        history = json.loads(_HISTORY_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}
    cutoff = (datetime.now(tz=timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    recent = [e for e in history if e.get("date", "") >= cutoff]
    result: dict[str, list[dict]] = {}
    for entry in recent:
        sh = entry.get("service_health", {})
        for name, status in sh.items():
            result.setdefault(name, []).append({"date": entry["date"], "status": status})
    return result

# ── Embedded cursor-api-proxy (Node / npx) ────────────────────────────────
_cursor_proxy_lock = threading.Lock()
_cursor_proxy_proc: subprocess.Popen | None = None


def _cursor_proxy_autostart_enabled(cfg: dict) -> bool:
    """If False, the app does not spawn npx cursor-api-proxy (user runs it manually). Default True."""
    ai = cfg.get("openai") or {}
    return ai.get("cursor_proxy_autostart", True) is not False


def _cursor_proxy_should_run(cfg: dict) -> bool:
    ai = cfg.get("openai") or {}
    if ai.get("provider") != "cursor":
        return False
    if not _cursor_proxy_autostart_enabled(cfg):
        return False
    key = (ai.get("api_key") or "").strip()
    if not key or key.lower() == "unused":
        return False
    return True


def _cursor_listen_host_port(ai: dict, cfg: dict) -> tuple[str, int]:
    """Host/port for CURSOR_BRIDGE_* from base_url (default 127.0.0.1:8765)."""
    base = (ai.get("base_url") or "").strip()
    if not base:
        return "127.0.0.1", 8765
    u = urlparse(base)
    host = u.hostname or "127.0.0.1"
    port = u.port if u.port is not None else 8765
    return host, port


def _cursor_health_url(host: str, port: int) -> str:
    h = "127.0.0.1" if host in ("0.0.0.0", "::", "") else host
    return f"http://{h}:{port}/health"


def _stop_cursor_proxy_unlocked() -> None:
    """Terminate embedded cursor-api-proxy (kill process tree on Windows)."""
    global _cursor_proxy_proc
    if _cursor_proxy_proc is None:
        return
    pid = _cursor_proxy_proc.pid
    proc = _cursor_proxy_proc
    _cursor_proxy_proc = None
    try:
        if sys.platform == "win32":
            cr = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(pid)],
                capture_output=True,
                timeout=20,
                creationflags=cr,
            )
        else:
            proc.terminate()
            try:
                proc.wait(timeout=10)
            except subprocess.TimeoutExpired:
                proc.kill()
    except Exception as exc:
        logger.warning("cursor proxy stop: %s", exc)
        try:
            proc.kill()
        except Exception:
            pass


def _find_npx_executable() -> str | None:
    """npx on PATH, or common Windows install paths (service may lack user PATH)."""
    w = shutil.which("npx") or shutil.which("npx.cmd")
    if w:
        return w
    if sys.platform == "win32":
        for p in (
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "nodejs" / "npx.cmd",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"))
            / "nodejs"
            / "npx.cmd",
            Path(os.environ.get("LocalAppData", "")) / "Programs" / "nodejs" / "npx.cmd",
        ):
            if p.is_file():
                return str(p)
    return None


def _nodejs_install_dirs() -> list[Path]:
    """Typical folders that contain node.exe + npx.cmd (Windows)."""
    if sys.platform != "win32":
        return []
    out: list[Path] = []
    for base in (
        Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "nodejs",
        Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "nodejs",
        Path(os.environ.get("LocalAppData", "")) / "Programs" / "nodejs",
    ):
        if (base / "node.exe").is_file():
            out.append(base.resolve())
    return out


def _prepend_nodejs_to_env(env: dict[str, str], npx_executable: str) -> None:
    """
    Ensure node.exe is on PATH for the child process.
    npx.cmd spawns `node`; embedded/web-server environments often lack the Node directory in PATH.
    """
    dirs: list[str] = []
    npx_p = Path(npx_executable).resolve()
    if npx_p.parent.is_dir() and (npx_p.parent / "node.exe").is_file():
        dirs.append(str(npx_p.parent))
    node = shutil.which("node") or (shutil.which("node.exe") if sys.platform == "win32" else None)
    if node:
        dirs.append(str(Path(node).resolve().parent))
    for d in _nodejs_install_dirs():
        s = str(d)
        if s not in dirs:
            dirs.append(s)
    # Dedupe, keep order
    seen: set[str] = set()
    ordered: list[str] = []
    for d in dirs:
        if d and d not in seen:
            seen.add(d)
            ordered.append(d)
    if not ordered:
        return
    sep = ";" if sys.platform == "win32" else ":"
    env["PATH"] = sep.join(ordered) + sep + env.get("PATH", "")


def _resolve_cursor_agent_from_config(cfg: dict) -> str | None:
    """Optional openai.cursor_agent_bin (file or folder), else auto-detect."""
    ai = cfg.get("openai") or {}
    manual = (ai.get("cursor_agent_bin") or "").strip()
    if manual:
        mp = Path(manual)
        if mp.is_file():
            return str(mp.resolve())
        if mp.is_dir():
            for name in ("agent.cmd", "agent.exe", "agent"):
                c = mp / name
                if c.is_file():
                    return str(c.resolve())
    return _find_cursor_agent_executable()


def _resolve_cursor_agent_cached(cfg: dict) -> str | None:
    """Same as _resolve_cursor_agent_from_config, cached until config YAML mtime changes."""
    global _cursor_agent_resolve_cache
    cpath = _config_yaml_path()
    try:
        mtime = cpath.stat().st_mtime
    except OSError:
        mtime = 0.0
    if _cursor_agent_resolve_cache is not None and _cursor_agent_resolve_cache[0] == mtime:
        return _cursor_agent_resolve_cache[1]
    resolved = _resolve_cursor_agent_from_config(cfg)
    _cursor_agent_resolve_cache = (mtime, resolved)
    return resolved


def _apply_cursor_agent_env(env: dict[str, str], agent_path: str) -> None:
    """
    cursor-api-proxy reads CURSOR_AGENT_* from process.env.
    On Windows, node.exe + index.js next to agent.cmd is the reliable invocation path
    (see cursor-api-proxy resolveAgentCommand).
    """
    p = Path(agent_path).resolve()
    env["CURSOR_AGENT_BIN"] = str(p)
    parent = p.parent
    if sys.platform == "win32":
        node_exe = parent / "node.exe"
        index_js = parent / "index.js"
        if node_exe.is_file() and index_js.is_file():
            env["CURSOR_AGENT_NODE"] = str(node_exe)
            env["CURSOR_AGENT_SCRIPT"] = str(index_js)
    sep = ";" if sys.platform == "win32" else ":"
    env["PATH"] = str(parent) + sep + env.get("PATH", "")


# Dirs to skip when walking for Cursor Agent CLI bundles (noise / huge trees).
_CURSOR_AGENT_WALK_SKIP_DIRS = frozenset(
    {
        "extensions",
        "CachedData",
        "CachedExtensionVSIXs",
        "logs",
        "WebStorage",
        "Crashpad",
        "GPUCache",
        "Code Cache",
        "node_modules",
        "site-packages",
        "Lib",
        "Miniconda3",
        "Anaconda3",
        ".git",
        "WindowsApps",
    }
)


def _walk_windows_agent_cmd_bundle(root: Path, max_depth: int) -> str | None:
    """Find agent.cmd with node.exe + index.js in the same directory under root."""
    if not root.is_dir():
        return None
    try:
        root = root.resolve()
    except OSError:
        return None
    try:
        for dirpath, dirnames, filenames in os.walk(str(root), topdown=True):
            rel = Path(dirpath)
            try:
                depth = len(rel.relative_to(root).parts)
            except ValueError:
                depth = 0
            if depth > max_depth:
                dirnames[:] = []
                continue
            dirnames[:] = [
                d
                for d in dirnames
                if d not in _CURSOR_AGENT_WALK_SKIP_DIRS
                and not d.startswith(".")
            ]
            if "agent.cmd" not in filenames:
                continue
            p = Path(dirpath) / "agent.cmd"
            d = p.parent
            if (d / "node.exe").is_file() and (d / "index.js").is_file():
                return str(p.resolve())
    except OSError:
        return None
    return None


def _iter_windows_agent_search_roots() -> list[tuple[Path, int]]:
    """
    (directory, max_depth) pairs — likely install locations for Cursor Agent CLI on Windows.
    Deeper search in ~/.cursor and name-*cursor* folders; shallower under generic Program Files trees.
    """
    out: list[tuple[Path, int]] = []
    seen: set[str] = set()

    def add(p: Path, depth: int) -> None:
        try:
            key = str(p.resolve())
        except OSError:
            key = str(p)
        if key in seen:
            return
        seen.add(key)
        out.append((p, depth))

    home = Path.home()
    la = os.environ.get("LOCALAPPDATA", "") or ""
    ad = os.environ.get("APPDATA", "") or ""
    pf = os.environ.get("ProgramFiles", r"C:\Program Files")
    pf86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")

    for p, d in (
        (home / ".cursor", 8),
        (Path(la) / "cursor-agent", 6),
        (home / "cursor", 6),
        (home / "AppData" / "Local" / "cursor-agent", 6),
        (Path(pf) / "cursor-agent", 5),
        (Path(pf) / "Cursor Agent", 5),
        (Path(pf86) / "cursor-agent", 5),
        (Path(pf86) / "Cursor Agent", 5),
        (Path(ad) / "Cursor", 5),
        (Path(ad) / "cursor-agent", 5),
    ):
        add(p, d)

    lp = Path(la) / "Programs"
    if lp.is_dir():
        try:
            for child in sorted(lp.iterdir(), key=lambda x: x.name.lower()):
                if not child.is_dir():
                    continue
                n = child.name.lower()
                depth = 7 if ("cursor" in n or "agent" in n) else 4
                add(child, depth)
        except OSError:
            pass

    # Optional extra drives: only small, conventional roots (not whole Program Files).
    for drive_letter in "DEFGHIJKLMNOPQRSTUVWXYZ":
        root_drive = Path(f"{drive_letter}:\\")
        if not root_drive.is_dir():
            continue
        for sub, depth in (
            ("tools", 6),
            ("opt", 6),
            ("Programs", 6),
        ):
            p = root_drive / sub
            if p.is_dir():
                add(p, depth)

    return out


def _find_windows_agent_bundle_cmd() -> str | None:
    """
    Cursor Agent CLI is separate from the IDE. Discover agent.cmd + node.exe + index.js
    by scanning common Windows install and user data locations.
    """
    if sys.platform != "win32":
        return None
    for base, depth in _iter_windows_agent_search_roots():
        found = _walk_windows_agent_cmd_bundle(base, depth)
        if found:
            return found
    return None


def _walk_unix_agent_bundle() -> str | None:
    """Non-Windows: find executable `agent` with index.js alongside (typical tarball layout)."""
    if sys.platform == "win32":
        return None
    roots: list[tuple[Path, int]] = []
    home = Path.home()
    for p, d in (
        (home / ".cursor", 8),
        (home / ".local" / "bin", 2),
        (home / ".local" / "share" / "cursor-agent", 6),
        (Path("/opt/cursor-agent"), 5),
        (Path("/usr/local/lib/cursor-agent"), 5),
    ):
        roots.append((p, d))
    xdg = os.environ.get("XDG_DATA_HOME", "")
    if xdg:
        roots.append((Path(xdg) / "cursor-agent", 6))
    skip = _CURSOR_AGENT_WALK_SKIP_DIRS
    for root, max_depth in roots:
        if not root.is_dir():
            continue
        try:
            r = root.resolve()
        except OSError:
            continue
        try:
            for dirpath, dirnames, filenames in os.walk(str(r), topdown=True):
                rel = Path(dirpath)
                try:
                    depth = len(rel.relative_to(r).parts)
                except ValueError:
                    depth = 0
                if depth > max_depth:
                    dirnames[:] = []
                    continue
                dirnames[:] = [x for x in dirnames if x not in skip and not x.startswith(".")]
                if "agent" not in filenames or "index.js" not in filenames:
                    continue
                agent_p = Path(dirpath) / "agent"
                if not agent_p.is_file():
                    continue
                try:
                    if os.access(agent_p, os.X_OK):
                        return str(agent_p.resolve())
                except OSError:
                    continue
        except OSError:
            continue
    return None


def _find_cursor_agent_executable() -> str | None:
    """
    Cursor CLI binary used by cursor-api-proxy for completions (not the same as Node).
    See: https://cursor.com/install — `agent` must be on PATH or CURSOR_AGENT_BIN.
    """
    for name in ("agent", "agent.cmd", "agent.exe"):
        w = shutil.which(name)
        if w:
            logger.info("Cursor Agent CLI (auto): %s (PATH)", w)
            return w
    if sys.platform == "win32":
        for root in (
            Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "cursor",
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Cursor",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Cursor",
        ):
            if not root.is_dir():
                continue
            for rel in (
                "resources/app/bin/agent.cmd",
                "resources/app/bin/agent.exe",
                "resources/app/bin/agent",
            ):
                p = (root / rel).resolve()
                if p.is_file():
                    logger.info("Cursor Agent CLI (auto): %s (IDE layout)", p)
                    return str(p)
        cur = Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "cursor"
        if cur.is_dir():
            try:
                for p in cur.rglob("agent.exe"):
                    if p.is_file():
                        logger.info("Cursor Agent CLI (auto): %s (rglob agent.exe)", p)
                        return str(p)
            except OSError:
                pass
        bundled = _find_windows_agent_bundle_cmd()
        if bundled:
            logger.info("Cursor Agent CLI (auto): %s", bundled)
            return bundled
    else:
        unix_agent = _walk_unix_agent_bundle()
        if unix_agent:
            logger.info("Cursor Agent CLI (auto): %s", unix_agent)
            return unix_agent
    return None


def _cursor_proxy_is_missing_node_npx(message: str) -> bool:
    """True when embedded proxy failed only because Node/npx is not installed or not in PATH."""
    low = (message or "").lower()
    if "npx" not in low and "node" not in low:
        return False
    return any(
        s in low
        for s in (
            "не найден",
            "not found",
            "no such file",
            "cannot find",
            "is not recognized",
        )
    )


def _cursor_proxy_log_path(cfg: dict) -> Path:
    dd = (cfg.get("general") or {}).get("data_dir", "data")
    return _REPO_ROOT / str(dd) / "cursor_proxy.log"


def _start_cursor_proxy_unlocked(cfg: dict) -> tuple[bool, str]:
    """Start npx cursor-api-proxy with CURSOR_API_KEY from config. Caller holds lock."""
    global _cursor_proxy_proc
    ai = cfg.get("openai") or {}
    key = (ai.get("api_key") or "").strip()
    host, port = _cursor_listen_host_port(ai, cfg)

    npx = _find_npx_executable()
    if not npx:
        return (
            False,
            "Cursor proxy не запущен: не найден npx (нужен Node.js: https://nodejs.org — поставьте LTS и убедитесь, что служба/консоль "
            "видит PATH, либо снимите галочку автозапуска proxy и запустите npx вручную).",
        )

    _stop_cursor_proxy_unlocked()

    log_path = _cursor_proxy_log_path(cfg)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    env = os.environ.copy()
    _prepend_nodejs_to_env(env, npx)
    agent_path = _resolve_cursor_agent_cached(cfg)
    if agent_path:
        _apply_cursor_agent_env(env, agent_path)
    else:
        logger.info(
            "Cursor Agent (CLI) не найден — чат Cursor в дашборде не будет работать, пока не установите CLI "
            "или не зададите openai.cursor_agent_bin. См. https://cursor.com/docs/cli/overview"
        )
    env["CURSOR_API_KEY"] = key
    env["CURSOR_BRIDGE_HOST"] = host
    env["CURSOR_BRIDGE_PORT"] = str(port)

    cr = 0
    if sys.platform == "win32" and hasattr(subprocess, "CREATE_NO_WINDOW"):
        cr = subprocess.CREATE_NO_WINDOW

    try:
        logf = open(log_path, "ab", buffering=0)
    except OSError as exc:
        return False, f"Cursor proxy: не удалось открыть лог {log_path}: {exc}"

    try:
        _cursor_proxy_proc = subprocess.Popen(
            [npx, "-y", "cursor-api-proxy"],
            env=env,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=logf,
            creationflags=cr,
        )
    except Exception as exc:
        try:
            logf.close()
        except Exception:
            pass
        return False, f"Cursor proxy: не удалось запустить npx: {exc}"

    # Wait for /health (proxy may download npm package on first run)
    health = _cursor_health_url(host, port)
    ok_health = False
    for _ in range(40):
        if _cursor_proxy_proc.poll() is not None:
            break
        try:
            r = httpx.get(health, timeout=1.5)
            if r.status_code < 500:
                ok_health = True
                break
        except Exception:
            pass
        time.sleep(0.5)

    if _cursor_proxy_proc.poll() is not None:
        return (
            False,
            f"Cursor proxy завершился сразу (см. лог {log_path}).",
        )
    if not ok_health:
        return (
            True,
            f"Cursor proxy запущен (PID {_cursor_proxy_proc.pid}), но /health пока не ответил — "
            f"первый запуск npx может долго качать пакет. Лог: {log_path}",
        )
    return True, f"Cursor proxy запущен (PID {_cursor_proxy_proc.pid}), {health} OK."


def _shutdown_embedded_cursor_proxy() -> None:
    with _cursor_proxy_lock:
        _stop_cursor_proxy_unlocked()


def sync_cursor_proxy_from_config(cfg: dict) -> dict:
    """Start or stop embedded cursor-api-proxy according to config. Thread-safe."""
    with _cursor_proxy_lock:
        if not _cursor_proxy_should_run(cfg):
            was_running = _cursor_proxy_running()
            _stop_cursor_proxy_unlocked()
            msg = ""
            if was_running:
                msg = "Встроенный Cursor proxy остановлен."
                if (
                    (cfg.get("openai") or {}).get("provider") == "cursor"
                    and not _cursor_proxy_autostart_enabled(cfg)
                ):
                    msg = "Встроенный Cursor proxy остановлен (автозапуск выключен)."
            return {
                "managed": True,
                "running": False,
                "ok": True,
                "message": msg,
            }
        ok, msg = _start_cursor_proxy_unlocked(cfg)
        running = _cursor_proxy_proc is not None and (
            _cursor_proxy_proc.poll() is None
        )
        warn = (not ok) and (not running) and _cursor_proxy_is_missing_node_npx(msg)
        return {
            "managed": True,
            "running": running,
            "ok": ok,
            "warning": warn,
            "message": msg,
        }


def _cursor_proxy_running() -> bool:
    return (
        _cursor_proxy_proc is not None and _cursor_proxy_proc.poll() is None
    )


# ── Rate limiting (action endpoints) ─────────────────────────────────────
_rate_limit_store: dict[str, float] = {}
_RATE_LIMIT_SECONDS = 15


def _check_rate_limit(key: str, window: float = _RATE_LIMIT_SECONDS) -> None:
    """Raise 429 if the same action key was invoked within *window* seconds."""
    now = time.monotonic()
    last = _rate_limit_store.get(key, 0.0)
    if now - last < window:
        wait = window - (now - last)
        raise HTTPException(429, f"Rate limit: try again in {wait:.1f}s")
    _rate_limit_store[key] = now


# ── State-change notifications ────────────────────────────────────────────
_notifications: list[dict] = []          # ring-buffer, newest last
_prev_build_statuses: dict[str, str] = {}
_prev_svc_statuses: dict[str, str] = {}
_prev_incident_active: bool = False
_prev_incident_sig: tuple[int, int, int, bool] = (0, 0, 0, False)  # (failed_builds, failed_tests, down_svcs, has_critical)
_NOTIFY_MAX = 200
_notify_id_seq = 0


def _event_feed_slim(entry: dict) -> dict:
    """Compact record for disk (matches in-app notification shape)."""
    out: dict = {
        "id": entry.get("id"),
        "ts": entry.get("ts"),
        "kind": entry.get("kind"),
        "level": entry.get("level"),
        "title": entry.get("title"),
        "detail": entry.get("detail"),
    }
    if entry.get("url"):
        out["url"] = entry["url"]
    if entry.get("critical"):
        out["critical"] = True
    return out


def _event_feed_append(entries: list[dict]) -> None:
    """Append state-change events to data/event_feed.json (capped)."""
    if not entries:
        return
    try:
        _EVENT_FEED_FILE.parent.mkdir(parents=True, exist_ok=True)
        cur: list = []
        if _EVENT_FEED_FILE.exists():
            raw = _EVENT_FEED_FILE.read_text(encoding="utf-8").strip()
            if raw:
                cur = json.loads(raw)
            if not isinstance(cur, list):
                cur = []
        for e in entries:
            cur.append(_event_feed_slim(e))
        if len(cur) > _EVENT_FEED_MAX:
            cur = cur[-_EVENT_FEED_MAX :]
        _EVENT_FEED_FILE.write_text(json.dumps(cur, ensure_ascii=False), encoding="utf-8")
    except Exception as exc:
        logger.warning("event_feed append failed: %s", exc)


def _event_feed_load(limit: int = 300) -> list[dict]:
    """Read newest *limit* persisted events (already chronological in file)."""
    if not _EVENT_FEED_FILE.exists():
        return []
    try:
        cur = json.loads(_EVENT_FEED_FILE.read_text(encoding="utf-8"))
        if not isinstance(cur, list):
            return []
        return cur[-limit:] if limit > 0 else cur
    except Exception as exc:
        logger.debug("event_feed load failed: %s", exc)
        return []


def _detect_state_changes(snapshot: "CISnapshot") -> None:
    """Diff current snapshot vs previous; append entries to _notifications."""
    global _prev_build_statuses, _prev_svc_statuses, _prev_incident_active, _prev_incident_sig, _notify_id_seq
    now_iso = datetime.now(tz=timezone.utc).isoformat()
    fail_st = {"failure", "unstable"}
    ok_st = {"success"}

    # Latest build per job (snapshot.builds newest-first)
    latest: dict[str, object] = {}
    for b in reversed(snapshot.builds):
        latest[b.job_name] = b

    for job_name, b in latest.items():
        prev = _prev_build_statuses.get(job_name)
        curr = b.status if isinstance(b.status, str) else b.status.value
        if prev is not None and prev != curr:
            if curr in fail_st and prev in ok_st:
                _notify_id_seq += 1
                ev = {
                    "id": _notify_id_seq,
                    "ts": now_iso,
                    "kind": "build_fail",
                    "level": "error",
                    "title": f"Job FAILED: {job_name}",
                    "detail": f"Status changed {prev} → {curr}",
                    "url": b.url,
                    "critical": b.critical,
                }
                _notifications.append(ev)
                _event_feed_append([ev])
            elif curr in ok_st and prev in fail_st:
                _notify_id_seq += 1
                ev = {
                    "id": _notify_id_seq,
                    "ts": now_iso,
                    "kind": "build_recovered",
                    "level": "ok",
                    "title": f"Job RECOVERED: {job_name}",
                    "detail": f"Status changed {prev} → {curr}",
                    "url": b.url,
                    "critical": b.critical,
                }
                _notifications.append(ev)
                _event_feed_append([ev])
        _prev_build_statuses[job_name] = curr

    for svc in snapshot.services:
        prev = _prev_svc_statuses.get(svc.name)
        curr = svc.status
        if prev is not None and prev != curr:
            if curr == "down" and prev in ("up", "degraded"):
                _notify_id_seq += 1
                ev = {
                    "id": _notify_id_seq,
                    "ts": now_iso,
                    "kind": "svc_down",
                    "level": "error",
                    "title": f"Service DOWN: {svc.name}",
                    "detail": f"Was {prev}, now down. {svc.detail or ''}",
                }
                _notifications.append(ev)
                _event_feed_append([ev])
            elif curr == "up" and prev == "down":
                _notify_id_seq += 1
                ev = {
                    "id": _notify_id_seq,
                    "ts": now_iso,
                    "kind": "svc_recovered",
                    "level": "ok",
                    "title": f"Service UP: {svc.name}",
                    "detail": f"Recovered from {prev}",
                }
                _notifications.append(ev)
                _event_feed_append([ev])
        _prev_svc_statuses[svc.name] = curr

    # Incident (aggregate) notification: emit once when an incident first appears.
    try:
        failed_builds = sum(1 for b in snapshot.builds if getattr(b, "status_normalized", None) in fail_st)
        failed_tests = sum(1 for t in snapshot.tests if getattr(t, "status_normalized", None) in ("failed", "error"))
        down_svcs = sum(1 for s in snapshot.services if getattr(s, "status_normalized", None) == "down")
        has_critical = any(
            getattr(b, "critical", False) and getattr(b, "status_normalized", None) in fail_st
            for b in snapshot.builds
        )
        active = (failed_builds > 0) or (failed_tests > 0) or (down_svcs > 0)
        sig = (failed_builds, failed_tests, down_svcs, bool(has_critical))
        if active and not _prev_incident_active:
            _notify_id_seq += 1
            lvl = "error" if (down_svcs > 0 or has_critical) else "warn"
            ev = {
                "id": _notify_id_seq,
                "ts": now_iso,
                "kind": "incident",
                "level": lvl,
                "title": "Incident detected",
                "detail": f"Failed builds: {failed_builds}, failed tests: {failed_tests}, services down: {down_svcs}",
                "url": "/?tab=incidents",
                "critical": bool(has_critical) or (down_svcs > 0),
            }
            _notifications.append(ev)
            _event_feed_append([ev])
        _prev_incident_active = active
        _prev_incident_sig = sig
    except Exception:
        # Never block build/service notifications on incident aggregation.
        pass

    # Trim ring-buffer
    if len(_notifications) > _NOTIFY_MAX:
        del _notifications[:len(_notifications) - _NOTIFY_MAX]


def _run_collect_sync(cfg: dict, *, force_full: bool = False) -> None:
    """Full collection — runs in a thread-pool executor (blocking)."""
    from clients.jenkins_client import JenkinsClient
    from clients.gitlab_client import GitLabClient
    from parsers.pytest_parser import PytestXMLParser
    from parsers.allure_parser import AllureJsonParser
    from docker_monitor.monitor import DockerMonitor
    from web.db import get_collector_state_int, set_collector_state_int

    since = datetime.now(tz=timezone.utc) - timedelta(
        days=cfg.get("general", {}).get("default_lookback_days", 7)
    )
    now = datetime.now(tz=timezone.utc)
    if force_full:
        snapshot = CISnapshot(collected_at=now, collect_meta={}, tests=[])
    else:
        prev = _load_snapshot() or CISnapshot()
        # Shallow copy: do not mutate the object returned from _load_snapshot() cache.
        # Fresh `tests` each collect avoids appending duplicates across runs.
        snapshot = prev.model_copy(
            update={
                "tests": [],
                "collect_meta": {},
                "collected_at": now,
            }
        )
    snap_lock = threading.Lock()
    global _instance_health
    health: list[dict[str, Any]] = []

    def _append_tests_live(recs: list[TestRecord]) -> None:
        if not recs:
            return
        with snap_lock:
            snapshot.tests.extend(recs)
        _maybe_save_partial(snapshot)

    def _progress(phase: str, main: str, sub: str | None = None) -> None:
        # This runs in a worker thread; updates are best-effort for UI polling.
        _collect_state["phase"] = phase
        _collect_state["progress_main"] = main
        _collect_state["progress_sub"] = sub
        _collect_state["progress_counts"] = {
            "builds": len(snapshot.builds),
            "tests": len(snapshot.tests),
            "services": len(snapshot.services),
        }
        # Heuristic: mark obvious failures as warn to make log filtering useful.
        lvl = "info"
        s = (sub or "").lower()
        if " error" in s or "failed" in s or "exception" in s or "traceback" in s or "retry" in s:
            lvl = "warn"
        _push_collect_log(phase, main, sub, lvl)

    def _build_key(b: object) -> str:
        try:
            bn = getattr(b, "build_number", None)
            inst_l = getattr(b, "source_instance", None) or ""
            return f"{getattr(b,'source','')}|{inst_l}|{getattr(b,'job_name','')}|{bn}|{getattr(b,'url','') or ''}"
        except Exception:
            return str(id(b))

    def _merge_build_records(new_records: list) -> None:
        if not new_records:
            return
        existing = snapshot.builds or []
        existing_by_key = {_build_key(b): b for b in existing}
        for b in new_records:
            existing_by_key[_build_key(b)] = b
        merged = list(existing_by_key.values())
        # Sort newest-first when possible
        try:
            merged.sort(key=lambda x: getattr(x, "started_at", None) or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
        except Exception:
            pass
        snapshot.builds = merged

    for inst in cfg.get("jenkins_instances", []):
        if not inst.get("enabled", True):
            continue
        label = inst.get("name", inst.get("url", "Jenkins"))
        inst_key = _config_instance_label(inst, kind="jenkins")
        shared_discovered: list[str] = []
        n_console_jobs_parsed = 0
        n_allure_jobs_parsed = 0
        t0 = time.monotonic()
        try:
            verify_ssl = bool(inst.get("verify_ssl", True))
            _progress("jenkins_builds", f"Jenkins: {label}", "Preparing job list…")

            # show_all_limit_jobs:
            # - absent -> safe default (50)
            # - <= 0  -> unlimited (None)
            _raw_limit = inst.get("show_all_limit_jobs", 50)
            try:
                _raw_limit = int(_raw_limit)
            except Exception:
                _raw_limit = 50
            show_all_limit_jobs = None if (_raw_limit is not None and int(_raw_limit) <= 0) else int(_raw_limit)

            client = JenkinsClient(
                url=inst["url"],
                username=inst.get("username", ""),
                token=inst.get("token", ""),
                jobs=inst.get("jobs", []),
                timeout=15,
                show_all=inst.get("show_all_jobs", False),
                show_all_limit_jobs=(
                    show_all_limit_jobs
                    if inst.get("show_all_jobs", False) and not inst.get("jobs") and show_all_limit_jobs is not None
                    else None
                ),
                verify_ssl=verify_ssl,
                progress_cb=lambda msg: _progress("jenkins_builds", f"Jenkins: {label}", msg),
                source_instance=inst_key,
            )
            if inst.get("show_all_jobs", False):
                try:
                    shared_discovered = client.fetch_job_list() or []
                except Exception as exc:
                    logger.warning("Jenkins [%s] fetch_job_list failed: %s", label, exc)
                    shared_discovered = []
            # max_builds: 0 means "unlimited" (no explicit limit in Jenkins request).
            try:
                effective_max_builds = int(inst.get("max_builds", 10))
            except Exception:
                effective_max_builds = 10
            if inst.get("show_all_jobs", False) and not inst.get("jobs"):
                cap = int(inst.get("show_all_max_builds", 20) or 20)
                if cap > 0:
                    effective_max_builds = min(effective_max_builds, cap)
            # Fast path: when show_all_jobs is enabled, fetch the latest status for many jobs in ONE request.
            if inst.get("show_all_jobs", False):
                limit_jobs = show_all_limit_jobs if (inst.get("show_all_jobs", False) and not inst.get("jobs")) else None
                _progress(
                    "jenkins_builds",
                    f"Jenkins: {label}",
                    f"Fetching lastBuild (bulk)… (limit_jobs={limit_jobs or 'all'})",
                )
                bulk_builds = client.fetch_last_builds_bulk(
                    since=since,
                    limit_jobs=limit_jobs,
                    depth=int(inst.get("show_all_depth", 4) or 4),
                )
                _merge_build_records(bulk_builds)
                # Bulk returns only lastCompletedBuild per job; optionally pull recent history per job.
                hist_n = int(inst.get("show_all_history_builds", 0) or 0)
                hist_job_cap = int(inst.get("show_all_history_jobs_cap", 45) or 45)
                if hist_n > 0 and bulk_builds:
                    crit_by: dict[str, bool] = {}
                    for j in inst.get("jobs") or []:
                        jn = (j.get("name") or "").strip()
                        if jn:
                            crit_by[jn] = bool(j.get("critical", False))
                    _progress(
                        "jenkins_builds",
                        f"Jenkins: {label}",
                        f"Fetching build history (≤{hist_n} builds/job, up to {hist_job_cap} jobs)…",
                    )
                    seen_hist: set[str] = set()
                    n_hist = 0
                    for b in bulk_builds:
                        if n_hist >= max(1, hist_job_cap):
                            break
                        jn = getattr(b, "job_name", None) or ""
                        if not jn or jn in seen_hist:
                            continue
                        if getattr(b, "build_number", None) is None:
                            continue
                        seen_hist.add(jn)
                        n_hist += 1
                        short = jn.rsplit("/", 1)[-1]
                        crit = bool(crit_by.get(jn) or crit_by.get(short))
                        try:
                            extra_hist = client.fetch_builds_for_job(
                                jn, since=since, max_builds=hist_n, critical=crit
                            )
                            if extra_hist:
                                _merge_build_records(extra_hist)
                        except Exception as exc:
                            logger.debug("Jenkins history fetch for %s: %s", jn, exc)
                # If /api/json?tree=jobs[name] (fetch_job_list) failed but bulk succeeded,
                # derive a stable "discovered jobs" list from the bulk payload so console/allure
                # parsing can still work and meta doesn't show "~0 jobs in index".
                if inst.get("show_all_jobs", False) and (not shared_discovered) and bulk_builds:
                    try:
                        shared_discovered = [b.job_name for b in bulk_builds if getattr(b, "job_name", None)]
                    except Exception:
                        shared_discovered = []
                # Build a quick status index for later (console/allure selection).
                last_status_by_job: dict[str, str] = {}
                for b in bulk_builds:
                    try:
                        last_status_by_job[b.job_name] = b.status_normalized
                    except Exception:
                        pass
                # Persist "last seen build number" per job for incremental collection.
                if _SQLITE_AVAILABLE and not force_full:
                    try:
                        base = str(inst.get("url", "")).rstrip("/")
                        for b in bulk_builds:
                            if not b.build_number:
                                continue
                            k = f"jenkins|{base}|{b.job_name}"
                            prev = get_collector_state_int(k, 0)
                            if int(b.build_number) > int(prev):
                                set_collector_state_int(k, int(b.build_number))
                    except Exception:
                        pass
                # Fallback "tests" derived from build results:
                # Many Jenkins jobs run exactly one test (-k=<job>) and console output
                # doesn't contain a stable per-test record for "passed". To keep the Tests tab
                # and Top failures populated even when console/allure parsing yields 0 records,
                # we emit one synthetic TestRecord per job build here.
                try:
                    for b in bulk_builds:
                        st = b.status_normalized
                        if st not in ("success", "failure", "unstable", "aborted"):
                            continue
                        t_status = "passed" if st == "success" else "failed" if st in ("failure", "unstable") else "skipped"
                        snapshot.tests.append(TestRecord(
                            source="jenkins_build",
                            suite=b.job_name,
                            test_name=b.job_name,
                            status=t_status,
                            duration_seconds=b.duration_seconds,
                            failure_message=None,
                            timestamp=b.started_at,
                        ))
                except Exception:
                    pass
                # Single write after builds + synthetic tests so the file never briefly holds tests=[].
                _maybe_save_partial(snapshot, force=True)
            else:
                _progress("jenkins_builds", f"Jenkins: {label}", f"Fetching builds… (max_builds={effective_max_builds})")
                _merge_build_records(
                    client.fetch_builds(since=since, max_builds=effective_max_builds)
                )
            health.append({
                "name": label,
                "kind": "jenkins",
                "ok": True,
                "error": None,
                "latency_ms": int((time.monotonic() - t0) * 1000),
            })
        except Exception as exc:
            logger.error("Jenkins [%s] builds failed: %s", label, exc)
            _push_collect_log("jenkins_builds", f"Jenkins: {label}", f"builds failed: {exc}", "error")
            health.append({
                "name": label,
                "kind": "jenkins",
                "ok": False,
                "error": str(exc),
                "latency_ms": None,
            })

        if inst.get("parse_console", False):
            try:
                from parsers.jenkins_console_parser import JenkinsConsoleParser
                jobs_for_console = inst.get("jobs", []) or []
                # When show_all_jobs is enabled, parse consoles for discovered jobs too (not only explicit jobs).
                if inst.get("show_all_jobs", False):
                    # Discover a limited set to avoid parsing hundreds of jobs by default.
                    # console_jobs_limit: 0 -> unlimited, absent -> default 25
                    raw_limit = inst.get("console_jobs_limit", 25)
                    try:
                        limit = int(raw_limit)
                    except Exception:
                        limit = 25
                    discovered = shared_discovered
                    if discovered:
                        # Only parse consoles for jobs that have a completed status of interest.
                        # This keeps the "parse everything" mode fast and avoids wasting requests on running/unknown.
                        wanted = {"success", "failure", "unstable"}
                        filtered = [n for n in discovered if last_status_by_job.get(n) in wanted]
                        if filtered:
                            discovered = filtered
                        if limit <= 0:
                            discovered_sel = discovered
                        else:
                            discovered_sel = discovered[: max(1, limit)]

                        # Merge explicit jobs (preserve critical flag) + discovered (non-critical by default)
                        explicit_by_name = {
                            (j.get("name") or ""): j for j in (jobs_for_console or []) if (j.get("name") or "")
                        }
                        merged_names = list(explicit_by_name.keys())
                        for n in discovered_sel:
                            if n not in explicit_by_name:
                                merged_names.append(n)
                        jobs_for_console = [
                            {
                                "name": n,
                                "critical": bool(explicit_by_name.get(n, {}).get("critical", False)),
                                "parse_console": True,
                            }
                            for n in merged_names
                        ]
                        logger.info(
                            "Jenkins [%s] console: discovered %d jobs, parsing %d (limit=%s, explicit=%d)",
                            label,
                            len(discovered),
                            len(jobs_for_console),
                            "all" if limit <= 0 else str(limit),
                            len(explicit_by_name),
                        )
                    else:
                        logger.warning(
                            "Jenkins [%s] console: show_all_jobs on but no jobs discovered; skipping console parse",
                            label,
                        )
                n_console_jobs_parsed = len(jobs_for_console) if jobs_for_console else 0
                if jobs_for_console:
                    _progress(
                        "jenkins_console",
                        f"Jenkins: {label}",
                        f"Parsing console ({len(jobs_for_console)} job(s))…",
                    )
                console_parser = JenkinsConsoleParser(
                    url=inst["url"],
                    username=inst.get("username", ""),
                    token=inst.get("token", ""),
                    jobs=jobs_for_console,
                    max_builds=int(inst.get("console_builds", 5) or 0),
                    workers=int(inst.get("console_workers", 8) or 8),
                    verify_ssl=bool(inst.get("verify_ssl", True)),
                    retries=int(inst.get("console_retries", 3) or 3),
                    backoff_seconds=float(inst.get("console_backoff_seconds", 0.8) or 0.8),
                    records_cb=_append_tests_live,
                    progress_cb=lambda msg: _progress("jenkins_console", f"Jenkins: {label}", msg),
                    timing_cb=lambda d: _collect_slow.append({
                        "ts": datetime.now(tz=timezone.utc).isoformat(),
                        "level": "info",
                        "instance": label,
                        "kind": d.get("kind"),
                        "job": d.get("job"),
                        "build": d.get("build"),
                        "elapsed_ms": d.get("elapsed_ms"),
                    }),
                )
                # Parser will stream records via records_cb for live UI; still return list for non-stream uses.
                _ = console_parser.fetch_tests()
            except Exception as exc:
                logger.error("Jenkins [%s] console parse failed: %s", label, exc)
                _push_collect_log("jenkins_console", f"Jenkins: {label}", f"console parse failed: {exc}", "error")

        if inst.get("parse_allure", False):
            try:
                from parsers.jenkins_allure_parser import JenkinsAllureParser

                jobs_for_allure = inst.get("jobs", []) or []
                # When show_all_jobs is enabled, parse Allure for discovered jobs too (not only explicit jobs).
                if inst.get("show_all_jobs", False):
                    # allure_jobs_limit: 0 -> unlimited, absent -> default 25
                    raw_limit = inst.get("allure_jobs_limit", 25)
                    try:
                        limit = int(raw_limit)
                    except Exception:
                        limit = 25
                    discovered = shared_discovered
                    if discovered:
                        # Allure is heavy; only parse jobs that are currently failing/unstable.
                        wanted = {"failure", "unstable"}
                        filtered = [n for n in discovered if last_status_by_job.get(n) in wanted]
                        if filtered:
                            discovered = filtered
                        if limit <= 0:
                            discovered_sel = discovered
                        else:
                            discovered_sel = discovered[: max(1, limit)]

                        explicit_by_name = {
                            (j.get("name") or ""): j for j in (jobs_for_allure or []) if (j.get("name") or "")
                        }
                        merged_names = list(explicit_by_name.keys())
                        for n in discovered_sel:
                            if n not in explicit_by_name:
                                merged_names.append(n)
                        jobs_for_allure = [
                            {
                                "name": n,
                                "critical": bool(explicit_by_name.get(n, {}).get("critical", False)),
                                "parse_allure": True,
                            }
                            for n in merged_names
                        ]
                        logger.info(
                            "Jenkins [%s] allure: discovered %d jobs, parsing %d (limit=%s, explicit=%d)",
                            label,
                            len(discovered),
                            len(jobs_for_allure),
                            "all" if limit <= 0 else str(limit),
                            len(explicit_by_name),
                        )
                    else:
                        logger.warning(
                            "Jenkins [%s] allure: show_all_jobs on but no jobs discovered; skipping allure parse",
                            label,
                        )

                n_allure_jobs_parsed = len(jobs_for_allure) if jobs_for_allure else 0
                if jobs_for_allure:
                    _progress(
                        "jenkins_allure",
                        f"Jenkins: {label}",
                        f"Parsing Allure ({len(jobs_for_allure)} job(s))…",
                    )
                try:
                    _ab_raw = inst.get("allure_builds")
                    if _ab_raw is None:
                        _ab_raw = inst.get("console_builds", 5)
                    allure_max_builds = int(_ab_raw)
                except Exception:
                    allure_max_builds = 5
                allure_parser = JenkinsAllureParser(
                    url=inst["url"],
                    username=inst.get("username", ""),
                    token=inst.get("token", ""),
                    jobs=jobs_for_allure,
                    max_builds=allure_max_builds,
                    workers=int(inst.get("allure_workers", 6) or 6),
                    verify_ssl=bool(inst.get("verify_ssl", True)),
                    progress_cb=lambda msg: _progress("jenkins_allure", f"Jenkins: {label}", msg),
                    retries=int(inst.get("allure_retries", 3) or 3),
                    backoff_seconds=float(inst.get("allure_backoff_seconds", 0.8) or 0.8),
                    records_cb=_append_tests_live,
                    timing_cb=lambda d: _collect_slow.append({
                        "ts": datetime.now(tz=timezone.utc).isoformat(),
                        "level": "info",
                        "instance": label,
                        "kind": d.get("kind"),
                        "job": d.get("job"),
                        "build": d.get("build"),
                        "elapsed_ms": d.get("elapsed_ms"),
                    }),
                )
                _ = allure_parser.fetch_tests()
            except Exception as exc:
                logger.error("Jenkins [%s] allure parse failed: %s", label, exc)
                _push_collect_log("jenkins_allure", f"Jenkins: {label}", f"allure parse failed: {exc}", "error")

        jobs_index_size = len(shared_discovered) if inst.get("show_all_jobs") else len(inst.get("jobs") or [])
        snapshot.collect_meta[f"jenkins:{label}"] = {
            "jobs_indexed": jobs_index_size,
            "console_jobs_parsed": n_console_jobs_parsed,
            "allure_jobs_parsed": n_allure_jobs_parsed,
        }

    for inst in cfg.get("gitlab_instances", []):
        if not inst.get("enabled", True):
            continue
        label = inst.get("name", inst.get("url", "GitLab"))
        gl_key = _config_instance_label(inst, kind="gitlab")
        t0 = time.monotonic()
        try:
            _progress("gitlab", f"GitLab: {label}", "Fetching pipelines…")
            from clients.gitlab_client import GitLabClient
            client = GitLabClient(
                url=inst.get("url", "https://gitlab.com"),
                token=inst.get("token", ""),
                projects=inst.get("projects", []),
                show_all=inst.get("show_all_projects", False),
                verify_ssl=bool(inst.get("verify_ssl", True)),
                source_instance=gl_key,
            )
            _merge_build_records(
                client.fetch_builds(since=since, max_builds=inst.get("max_pipelines", 10))
            )
            health.append({
                "name": label,
                "kind": "gitlab",
                "ok": True,
                "error": None,
                "latency_ms": int((time.monotonic() - t0) * 1000),
            })
        except Exception as exc:
            logger.error("GitLab [%s] failed: %s", label, exc)
            health.append({
                "name": label,
                "kind": "gitlab",
                "ok": False,
                "error": str(exc),
                "latency_ms": None,
            })

    p_cfg = cfg.get("parsers", {})
    pytest_parser = PytestXMLParser()
    allure_parser = AllureJsonParser()
    for d in p_cfg.get("pytest_xml_dirs", []):
        try:
            snapshot.tests.extend(pytest_parser.parse_directory(d))
        except Exception as exc:
            logger.error("pytest parser failed for %s: %s", d, exc)
    for d in p_cfg.get("allure_json_dirs", []):
        try:
            snapshot.tests.extend(allure_parser.parse_directory(d))
        except Exception as exc:
            logger.error("allure parser failed for %s: %s", d, exc)

    dm_cfg = cfg.get("docker_monitor", {})
    if dm_cfg.get("enabled"):
        t0 = time.monotonic()
        try:
            _progress("docker", "Docker / HTTP", "Running checks…")
            monitor = DockerMonitor(
                containers=dm_cfg.get("containers", []),
                http_checks=dm_cfg.get("http_checks", []),
                timeout=dm_cfg.get("timeout_seconds", 5),
                show_all=dm_cfg.get("show_all_containers", False),
            )
            snapshot.services = monitor.check_all()
            health.append({
                "name": "Docker monitor",
                "kind": "docker",
                "ok": True,
                "error": None,
                "latency_ms": int((time.monotonic() - t0) * 1000),
            })
        except Exception as exc:
            logger.error("Docker monitor failed: %s", exc)
            health.append({
                "name": "Docker monitor",
                "kind": "docker",
                "ok": False,
                "error": str(exc),
                "latency_ms": None,
            })

    _instance_health = health
    save_snapshot(snapshot)
    _progress("done", "Collect finished", None)
    logger.info(
        "Auto-collect done: builds=%d tests=%d services=%d",
        len(snapshot.builds), len(snapshot.tests), len(snapshot.services),
    )


async def _do_collect(cfg: dict, *, force_full: bool = False) -> None:
    """Async wrapper: run collection in thread pool, update shared state."""
    if _collect_state["is_collecting"]:
        logger.info("Collection already in progress, skipping.")
        return
    _collect_state["is_collecting"] = True
    _collect_state["started_at"] = datetime.now(tz=timezone.utc).isoformat()
    _collect_state["phase"] = "starting"
    _collect_state["progress_main"] = "Starting collect…"
    _collect_state["progress_sub"] = None
    _collect_state["progress_counts"] = {"builds": 0, "tests": 0, "services": 0}
    _collect_state["last_error"] = None
    try:
        _collect_logs.clear()
        _collect_slow.clear()
    except Exception:
        pass
    _push_collect_log("starting", "Starting collect…", None, "info")
    try:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, lambda: _run_collect_sync(cfg, force_full=force_full))
        _collect_state["last_collected_at"] = datetime.now(tz=timezone.utc).isoformat()
    except Exception as exc:
        logger.error("Collection error: %s", exc)
        _collect_state["last_error"] = str(exc)
    finally:
        _collect_state["is_collecting"] = False
        _collect_state["started_at"] = None
        try:
            await _sse_broadcast_async(
                {
                    "type": "collect_done",
                    "last_collected_at": _collect_state.get("last_collected_at"),
                    "error": _collect_state.get("last_error"),
                    "revision": _data_revision,
                }
            )
        except Exception:
            pass


async def _collect_loop(cfg: dict) -> None:
    """Collect immediately on start, then repeat every interval."""
    while True:
        # Re-check settings each cycle (interval can change, and the loop can be disabled).
        if not _auto_collect_enabled:
            await asyncio.sleep(1.0)
            continue
        interval = int(_collect_state.get("interval_seconds") or 300)
        await _do_collect(cfg, force_full=False)
        await asyncio.sleep(max(5, interval))


# ── FastAPI lifespan ──────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    global _collect_task, _main_loop
    _main_loop = asyncio.get_running_loop()
    cfg = _load_yaml_config()
    w_cfg = cfg.get("web", {})

    # Initialize SQLite persistent storage
    if _SQLITE_AVAILABLE:
        data_dir = Path(cfg.get("general", {}).get("data_dir", "data"))
        try:
            init_db(data_dir)
            logger.info("SQLite history DB initialized at %s", data_dir / "monitor.db")
        except Exception as exc:
            logger.warning("SQLite init failed (non-fatal): %s", exc)

    # Background auto-collect is enabled only when UI LIVE mode is enabled (toggled from the client).
    # We still create the task so it can be enabled later without restarting the server.
    interval = int(w_cfg.get("collect_interval_seconds", 300))
    _collect_state["interval_seconds"] = interval
    if w_cfg.get("auto_collect", True):
        logger.info("Auto-collect is configured (interval=%ds) but starts only when LIVE is enabled.", interval)
        _collect_task = asyncio.create_task(_collect_loop(cfg))
    else:
        logger.info("Auto-collect disabled in config (web.auto_collect=false).")
    proxy_paths = sorted(
        {getattr(r, "path", "") for r in app.routes if getattr(r, "path", "") and "proxy" in getattr(r, "path", "")}
    )
    logger.info(
        "Web app build=%s config=%s proxy_routes=%s",
        _APP_BUILD,
        _config_yaml_path(),
        proxy_paths or "none",
    )
    try:
        cp = await asyncio.to_thread(sync_cursor_proxy_from_config, cfg)
        logger.info("Embedded Cursor proxy: %s", cp)
    except Exception as exc:
        logger.warning("Embedded Cursor proxy startup failed: %s", exc)
    yield
    try:
        await asyncio.to_thread(_shutdown_embedded_cursor_proxy)
    except Exception as exc:
        logger.warning("Embedded Cursor proxy shutdown: %s", exc)
    if _collect_task:
        _collect_task.cancel()
        try:
            await _collect_task
        except asyncio.CancelledError:
            pass


app = FastAPI(title="CI/CD Monitor", version="1.0.0", lifespan=lifespan)

_static_dir = Path(__file__).resolve().parent / "static"
if _static_dir.is_dir():
    app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")


@app.middleware("http")
async def add_request_id(request: Request, call_next):
    rid = (request.headers.get("x-request-id") or "").strip() or str(uuid.uuid4())
    request.state.request_id = rid
    response = await call_next(request)
    response.headers["X-Request-ID"] = rid
    return response


def _rid(request: Request | None) -> str:
    if request is None:
        return "-"
    return getattr(request.state, "request_id", "-")


def _config_instance_label(inst: dict[str, Any], *, kind: str) -> str:
    """Stable label for a jenkins_instances / gitlab_instances entry (merge + UI grouping)."""
    n = str(inst.get("name") or "").strip()
    if n:
        return n[:240]
    u = str(inst.get("url") or "").strip()
    if u:
        try:
            net = urlparse(u).netloc
            return (net or u.rstrip("/"))[:240]
        except Exception:
            return u[:240]
    return "Jenkins" if kind == "jenkins" else "GitLab"


def _enabled_ci_bases(cfg: dict[str, Any], kind: str) -> list[str]:
    insts = cfg.get(f"{kind}_instances", []) or []
    out: list[str] = []
    for inst in insts:
        if not inst.get("enabled", True):
            continue
        u = str(inst.get("url", "") or "").strip()
        if u:
            out.append(u.rstrip("/"))
    return out


def _build_url_matches_ci_bases(b: Any, bases: list[str]) -> bool:
    """Whether build.url belongs to one of the configured instance roots.

    Jenkins sometimes returns a *path-only* URL (``/job/...``); resolve it against
    each enabled base so those builds are not dropped by the dashboard filter.
    """
    if not bases:
        return False
    bu_raw = str(getattr(b, "url", None) or "").strip()
    if not bu_raw:
        return True
    bu = bu_raw.rstrip("/")
    bl = bu.lower()
    for base in bases:
        br = str(base).rstrip("/")
        if not br:
            continue
        brl = br.lower()
        if bu.startswith(br) or bl.startswith(brl):
            return True
        if bu_raw.startswith("/"):
            try:
                joined = urljoin(br + "/", bu_raw).rstrip("/").lower()
                if joined.startswith(brl):
                    return True
            except Exception:
                pass
    return False


def _is_snapshot_build_enabled(b: Any, cfg: dict[str, Any]) -> bool:
    try:
        src = (b.source or "").lower()
    except Exception:
        return True
    if src == "jenkins":
        bases = _enabled_ci_bases(cfg, "jenkins")
        return _build_url_matches_ci_bases(b, bases)
    if src == "gitlab":
        bases = _enabled_ci_bases(cfg, "gitlab")
        return _build_url_matches_ci_bases(b, bases)
    return True


def _inst_label_for_build_with_cfg(b: Any, cfg: dict[str, Any]) -> str | None:
    """Instance column / filter label: prefer ``source_instance``, else URL → config match."""
    try:
        stored = getattr(b, "source_instance", None)
    except Exception:
        stored = None
    if isinstance(stored, str) and stored.strip():
        return stored.strip()
    try:
        src = (b.source or "").lower()
    except Exception:
        return None
    try:
        bu = (b.url or "").rstrip("/")
    except Exception:
        bu = ""
    if src == "jenkins":
        for inst in (cfg.get("jenkins_instances", []) or []):
            if not inst.get("enabled", True):
                continue
            base = str(inst.get("url", "") or "").rstrip("/")
            if base and bu.startswith(base):
                return _config_instance_label(inst, kind="jenkins")
            if base and bu.startswith("/"):
                try:
                    joined = urljoin(base + "/", str(getattr(b, "url", None) or "").strip())
                    if joined.rstrip("/").startswith(base):
                        return _config_instance_label(inst, kind="jenkins")
                except Exception:
                    pass
        return None
    if src == "gitlab":
        for inst in (cfg.get("gitlab_instances", []) or []):
            if not inst.get("enabled", True):
                continue
            base = str(inst.get("url", "") or "").rstrip("/")
            if base and bu.startswith(base):
                return _config_instance_label(inst, kind="gitlab")
            if base and bu.startswith("/"):
                try:
                    joined = urljoin(base + "/", str(getattr(b, "url", None) or "").strip())
                    if joined.rstrip("/").startswith(base):
                        return _config_instance_label(inst, kind="gitlab")
                except Exception:
                    pass
        return None
    return None


# ── REST endpoints ────────────────────────────────────────────────────────

@app.get("/api/status", response_class=JSONResponse)
async def api_status():
    """Return the full snapshot as JSON."""
    snap = _load_snapshot()
    if snap is None:
        return JSONResponse(
            {"error": "No data yet. Run ci_monitor.py to collect data."},
            status_code=404,
        )
    cfg = _load_yaml_config()

    data = json.loads(snap.model_dump_json())
    builds = [b for b in (snap.builds or []) if _is_snapshot_build_enabled(b, cfg)]
    data["builds"] = [
        dict(json.loads(b.model_dump_json()), instance=_inst_label_for_build_with_cfg(b, cfg)) for b in builds
    ]
    return data


async def api_builds(
    page: int = 1,
    per_page: int = 20,
    source: str = "",
    instance: str = "",
    status: str = "",
    job: str = "",
    hours: int = 0,
):
    """
    Paginated build records.
    Query params: page, per_page, source (jenkins|gitlab|…), status (success|failure|…),
                  job (substring), hours (0=all, >0 filter by started_at recency)
    Returns: {items, page, per_page, total, has_more}
    """
    page = max(1, int(page or 1))
    per_page = min(max(1, int(per_page or 20)), 200)
    snap = await _load_snapshot_async()
    if snap is None:
        raise HTTPException(404, "No snapshot data found.")

    cfg = _load_yaml_config()

    items = [b for b in snap.builds if _is_snapshot_build_enabled(b, cfg)]
    if source:
        items = [b for b in items if b.source.lower() == source.lower()]
    if instance:
        want_inst = instance.strip().lower()
        if want_inst:
            items = [
                b
                for b in items
                if (_inst_label_for_build_with_cfg(b, cfg) or "").strip().lower() == want_inst
            ]
    if status:
        want = normalize_build_status(status)
        items = [b for b in items if b.status_normalized == want]
    if job:
        items = [b for b in items if job.lower() in b.job_name.lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [
            b for b in items
            if b.started_at and b.started_at.replace(tzinfo=timezone.utc if b.started_at.tzinfo is None else b.started_at.tzinfo) >= cutoff
        ]

    # Totals per source||instance for the full filtered list (UI group headers).
    group_counts: dict[str, dict[str, int]] = {}
    for b in items:
        gk = f"{(b.source or '').strip().lower()}||{(_inst_label_for_build_with_cfg(b, cfg) or '').strip().lower()}"
        if gk not in group_counts:
            group_counts[gk] = {"fail": 0, "warn": 0, "ok": 0, "total": 0}
        rec = group_counts[gk]
        rec["total"] += 1
        sn = b.status_normalized
        if sn == "failure":
            rec["fail"] += 1
        elif sn == "unstable":
            rec["warn"] += 1
        elif sn == "success":
            rec["ok"] += 1

    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page
    page_items = items[start:end]

    job_ctx = _job_build_analytics(snap)
    out_items = []
    for b in page_items:
        row = json.loads(b.model_dump_json())
        row["analytics"] = job_ctx.get(b.job_name, {})
        row["instance"] = _inst_label_for_build_with_cfg(b, cfg)
        out_items.append(row)

    return {
        "items": out_items,
        "page": page,
        "per_page": per_page,
        "total": total,
        "has_more": end < total,
        "group_counts": group_counts,
    }


async def api_instances():
    """Return enabled instance labels for filter dropdowns.

    Labels match ``_config_instance_label`` / ``source_instance`` on builds
    (explicit ``name`` in YAML, otherwise URL host), so instance filtering
    stays consistent even when ``name`` is omitted in config.
    """
    cfg = _load_yaml_config()
    out: list[dict[str, str]] = []
    for inst in (cfg.get("jenkins_instances", []) or []):
        if not inst.get("enabled", True):
            continue
        if not str(inst.get("url", "") or "").strip():
            continue
        out.append({"source": "jenkins", "name": _config_instance_label(inst, kind="jenkins")})
    for inst in (cfg.get("gitlab_instances", []) or []):
        if not inst.get("enabled", True):
            continue
        if not str(inst.get("url", "") or "").strip():
            continue
        out.append({"source": "gitlab", "name": _config_instance_label(inst, kind="gitlab")})
    out.sort(key=lambda x: (x.get("source", ""), x.get("name", "")))
    return out


async def api_builds_history(
    page: int = 1,
    per_page: int = 50,
    job: str = "",
    source: str = "",
    status: str = "",
    days: int = 30,
):
    """
    Paginated build history from SQLite (across collects). Same shape idea as /api/builds
    but backed by monitor.db when available.
    """
    if not _SQLITE_AVAILABLE or _db_query_builds_history is None:
        return {
            "items": [],
            "page": page,
            "per_page": per_page,
            "total": 0,
            "has_more": False,
            "source": "none",
            "note": "sqlite_unavailable",
        }
    try:
        data = _db_query_builds_history(
            job=job,
            source=source,
            status=status,
            page=max(1, page),
            per_page=min(max(1, per_page), 200),
            days=min(max(1, days), 365),
        )
    except Exception as exc:
        logger.warning("api_builds_history: %s", exc)
        raise HTTPException(500, str(exc)) from exc
    data["page"] = max(1, page)
    data["per_page"] = min(max(1, per_page), 200)
    data["source"] = "sqlite"
    return data


def _dashboard_summary_payload() -> dict[str, Any]:
    cfg = _load_yaml_config()
    w_cfg = cfg.get("web", {})
    interval = int(w_cfg.get("collect_interval_seconds", 300))
    stale_threshold = interval * 2

    snap = _load_snapshot()
    collected_at: str | None = None
    age_seconds: float | None = None
    stale = False
    counts = {
        "builds": 0,
        "failed_builds": 0,
        "failed_tests": 0,
        "tests_total": 0,
        "services_down": 0,
    }
    if snap:
        counts["builds"] = len(snap.builds)
        counts["failed_builds"] = sum(
            1 for b in snap.builds if b.status_normalized in ("failure", "unstable")
        )
        counts["failed_tests"] = sum(
            1 for t in snap.tests if t.status_normalized in ("failed", "error")
        )
        counts["tests_total"] = len(snap.tests)
        counts["services_down"] = sum(1 for s in snap.services if s.status_normalized == "down")
        ca = snap.collected_at
        if ca.tzinfo is None:
            ca = ca.replace(tzinfo=timezone.utc)
        else:
            ca = ca.astimezone(timezone.utc)
        collected_at = ca.isoformat()
        age_seconds = (datetime.now(tz=timezone.utc) - ca).total_seconds()
        stale = age_seconds > stale_threshold

    partial_errors: list[dict[str, Any]] = []
    if _collect_state.get("last_error"):
        partial_errors.append({"source": "collect", "message": _collect_state["last_error"]})
    for h in _instance_health:
        if not h.get("ok"):
            partial_errors.append({
                "source": h.get("kind"),
                "name": h.get("name"),
                "message": h.get("error"),
            })

    return {
        "data_revision": _data_revision,
        "snapshot": {
            "collected_at": collected_at,
            "age_seconds": round(age_seconds, 1) if age_seconds is not None else None,
            "stale": stale,
            "stale_threshold_seconds": stale_threshold,
        },
        "counts": counts,
        "collect": {
            "is_collecting": _collect_state["is_collecting"],
            "last_collected_at": _collect_state["last_collected_at"],
            "last_error": _collect_state["last_error"],
            "interval_seconds": interval,
        },
        "partial_errors": partial_errors,
        "instance_health": list(_instance_health),
        "parse_coverage": (snap.collect_meta if snap else {}) or {},
    }


@app.get("/api/dashboard/summary", response_class=JSONResponse)
async def api_dashboard_summary():
    """Single payload for dashboard boot: counts, freshness, collect + source health."""
    return _dashboard_summary_payload()


async def api_instances_health():
    """Last collect: Jenkins / GitLab / Docker poll latency and errors."""
    return {
        "last_collected_at": _collect_state.get("last_collected_at"),
        "instances": list(_instance_health),
    }


@app.get("/api/meta", response_class=JSONResponse)
async def api_meta():
    """Dashboard metadata: snapshot freshness, collect state, correlation, job analytics."""
    cfg = _load_yaml_config()
    w_cfg = cfg.get("web", {})
    interval = int(w_cfg.get("collect_interval_seconds", 300))
    stale_threshold = interval * 2

    snap = await _load_snapshot_async()
    collected_at: str | None = None
    age_seconds: float | None = None
    stale = False
    job_analytics: dict = {}
    if snap:
        job_analytics = _job_build_analytics(snap)
        ca = snap.collected_at
        if ca.tzinfo is None:
            ca = ca.replace(tzinfo=timezone.utc)
        else:
            ca = ca.astimezone(timezone.utc)
        collected_at = ca.isoformat()
        age_seconds = (datetime.now(tz=timezone.utc) - ca).total_seconds()
        stale = age_seconds > stale_threshold

    return {
        "data_revision": _data_revision,
        "snapshot": {
            "collected_at": collected_at,
            "age_seconds": round(age_seconds, 1) if age_seconds is not None else None,
            "stale": stale,
            "stale_threshold_seconds": stale_threshold,
        },
        "collect": {
            "is_collecting": _collect_state["is_collecting"],
            "last_collected_at": _collect_state["last_collected_at"],
            "last_error": _collect_state["last_error"],
            "interval_seconds": interval,
        },
        "correlation": _correlation_last_hour(),
        "job_analytics": job_analytics,
        "parse_coverage": (snap.collect_meta if snap else {}) or {},
    }


@app.get("/api/stream/events")
async def sse_events(request: Request):
    """Server-Sent Events: collect_done + heartbeats. Clients fall back to polling if unavailable."""

    async def event_gen():
        q: asyncio.Queue = asyncio.Queue(maxsize=64)
        _sse_queues.add(q)
        try:
            yield f"data: {json.dumps({'type': 'hello', 'revision': _data_revision})}\n\n"
            while True:
                if await request.is_disconnected():
                    break
                try:
                    ev = await asyncio.wait_for(q.get(), timeout=25.0)
                    yield f"data: {json.dumps(ev)}\n\n"
                except asyncio.TimeoutError:
                    yield ": ping\n\n"
        finally:
            _sse_queues.discard(q)

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _aggregate_top_failing_tests(
    tests: list[Any],
    *,
    top_n: int,
    suite_sub: str = "",
    name_sub: str = "",
    message_max: int = 300,
) -> list[dict[str, Any]]:
    """Group failed/error runs by test_name; pick error text from the latest run that has one."""
    by_name: dict[str, list[Any]] = defaultdict(list)
    for t in tests:
        if t.status_normalized in ("failed", "error"):
            by_name[t.test_name].append(t)

    def _ts(rec: Any) -> datetime:
        ts = rec.timestamp
        if ts is None:
            return datetime.min.replace(tzinfo=timezone.utc)
        if ts.tzinfo is None:
            return ts.replace(tzinfo=timezone.utc)
        return ts.astimezone(timezone.utc)

    rows: list[dict[str, Any]] = []
    no_detail = "(no failure text in report)"
    for tname, recs in by_name.items():
        recs_sorted = sorted(recs, key=_ts, reverse=True)
        latest = recs_sorted[0]
        suite_val = latest.suite
        msg: str | None = None
        for r in recs_sorted:
            fm = r.failure_message
            if fm and str(fm).strip().lower() != "null":
                msg = str(fm).strip()
                break
        if not msg:
            msg = no_detail
        if message_max > 0 and len(msg) > message_max:
            msg = msg[:message_max]
        rows.append({
            "test_name": tname,
            "count": len(recs),
            "suite": suite_val,
            "message": msg,
        })

    rows.sort(key=lambda x: (-x["count"], x["test_name"]))
    rows = rows[:top_n]

    if suite_sub:
        sl = suite_sub.lower()
        rows = [r for r in rows if sl in (r.get("suite") or "").lower()]
    if name_sub:
        nl = name_sub.lower()
        rows = [r for r in rows if nl in r["test_name"].lower()]
    return rows


def _filter_tests_by_source(items: list[Any], source: str) -> list[Any]:
    s = (source or "").strip().lower()
    if not s:
        return items
    if s == "synthetic":
        return [t for t in items if (t.source or "").strip().lower() == "jenkins_build"]
    if s == "real":
        return [t for t in items if (t.source or "").strip().lower() != "jenkins_build"]
    return [t for t in items if (t.source or "").strip().lower() == s]


def _filter_tests_by_lookback_hours(
    tests: list[Any],
    *,
    hours: int = 0,
    days: int = 0,
) -> list[Any]:
    """Keep tests whose timestamp falls within the lookback window (UTC). days overrides hours if both set."""
    lookback_h = 0
    if days and int(days) > 0:
        lookback_h = int(days) * 24
    elif hours and int(hours) > 0:
        lookback_h = int(hours)
    if lookback_h <= 0:
        return list(tests)

    cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=lookback_h)

    def _ts(rec: Any) -> datetime:
        ts = getattr(rec, "timestamp", None)
        if ts is None:
            return datetime.min.replace(tzinfo=timezone.utc)
        if ts.tzinfo is None:
            return ts.replace(tzinfo=timezone.utc)
        return ts.astimezone(timezone.utc)

    return [t for t in tests if _ts(t) >= cutoff]


def _tests_breakdown_real_vs_synth(items: list[Any]) -> dict[str, int]:
    real_total = 0
    real_failed = 0
    syn_total = 0
    syn_failed = 0
    for t in items:
        src = (t.source or "").strip().lower()
        is_syn = (src == "jenkins_build")
        if is_syn:
            syn_total += 1
            if t.status_normalized in ("failed", "error"):
                syn_failed += 1
        else:
            real_total += 1
            if t.status_normalized in ("failed", "error"):
                real_failed += 1
    return {
        "real_total": real_total,
        "real_failed": real_failed,
        "synthetic_total": syn_total,
        "synthetic_failed": syn_failed,
    }

async def api_tests(
    page: int = 1,
    per_page: int = 30,
    status: str = "",
    suite: str = "",
    name: str = "",
    hours: int = 0,
    source: str = "",
):
    """
    Paginated test records (all individual test runs, not aggregated).
    Query params: page, per_page, status (passed|failed|skipped|error), suite, name (substring),
                  hours (0=all, >0 filter by timestamp recency)
    Returns: {items, page, per_page, total, has_more}
    """
    snap = await _load_snapshot_async()
    if snap is None:
        raise HTTPException(404, "No snapshot data found.")

    items = snap.tests
    if status:
        want = normalize_test_status(status)
        items = [t for t in items if t.status_normalized == want]
    if suite:
        items = [t for t in items if suite.lower() in (t.suite or "").lower()]
    if name:
        items = [t for t in items if name.lower() in t.test_name.lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [
            t for t in items
            if t.timestamp and t.timestamp.replace(tzinfo=timezone.utc if t.timestamp.tzinfo is None else t.timestamp.tzinfo) >= cutoff
        ]

    # Breakdown should be "honest" and not depend on source/status selection.
    # So we compute it before applying source/status filters.
    breakdown_base = snap.tests
    if suite:
        breakdown_base = [t for t in breakdown_base if suite.lower() in (t.suite or "").lower()]
    if name:
        breakdown_base = [t for t in breakdown_base if name.lower() in t.test_name.lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        breakdown_base = [
            t for t in breakdown_base
            if t.timestamp and t.timestamp.replace(tzinfo=timezone.utc if t.timestamp.tzinfo is None else t.timestamp.tzinfo) >= cutoff
        ]
    breakdown = _tests_breakdown_real_vs_synth(breakdown_base)

    # Now apply source filter to items for actual table content.
    items = _filter_tests_by_source(items, source)

    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page
    page_items = items[start:end]

    return {
        "items": [json.loads(t.model_dump_json()) for t in page_items],
        "page": page,
        "per_page": per_page,
        "total": total,
        "has_more": end < total,
        "breakdown": breakdown,
    }


async def api_top_failures(
    n: int = 50,
    page: int = 1,
    per_page: int = 20,
    suite: str = "",
    name: str = "",
    source: str = "",
    hours: int = 0,
    days: int = 0,
):
    """
    Aggregated top-failing tests, paginated.
    Query: hours or days (days wins if both >0) limit by test timestamp within the current snapshot.
    """
    snap = _load_snapshot()
    if snap is None:
        raise HTTPException(404, "No snapshot data found.")

    tests_items = _filter_tests_by_lookback_hours(snap.tests, hours=int(hours or 0), days=int(days or 0))

    src = (source or "").strip().lower()
    # When source=all (empty), keep rows separated by parser source to avoid mixing
    # synthetic "job as test" with real test cases.
    if not src:
        counter: Counter = Counter()
        messages: dict[str, str] = {}
        msg_ts: dict[str, datetime] = {}
        suites: dict[str, str] = {}
        suite_ts: dict[str, datetime] = {}
        sources: dict[str, str] = {}

        def _rec_ts(rec: Any) -> datetime:
            ts = getattr(rec, "timestamp", None)
            if ts is None:
                return datetime.min.replace(tzinfo=timezone.utc)
            if ts.tzinfo is None:
                return ts.replace(tzinfo=timezone.utc)
            return ts.astimezone(timezone.utc)

        def _candidate_message(rec: Any) -> str | None:
            src_l = (rec.source or "").strip().lower()
            fm = (rec.failure_message or "")
            if fm and str(fm).strip().lower() != "null":
                return str(fm).strip()
            if src_l == "jenkins_build":
                st = (rec.status or "").strip().lower()
                return "Job failed/unstable" if st in ("failed", "error") else "Job failed"
            return None

        for t in tests_items:
            if t.status_normalized in ("failed", "error"):
                key = f"{(t.source or 'unknown')}::{t.test_name}"
                counter[key] += 1
                sources[key] = (t.source or "unknown")
                ts = _rec_ts(t)
                if t.suite and str(t.suite).strip():
                    prev = suite_ts.get(key)
                    if prev is None or ts >= prev:
                        suite_ts[key] = ts
                        suites[key] = str(t.suite).strip()

                cand = _candidate_message(t)
                if not cand:
                    continue
                prev_m_ts = msg_ts.get(key)
                if prev_m_ts is None or ts >= prev_m_ts:
                    msg_ts[key] = ts
                    messages[key] = cand[:300]

        no_detail = "(no failure text in report)"
        all_items = [
            {
                "source": sources.get(k),
                "test_name": k.split("::", 1)[1],
                "count": c,
                "suite": suites.get(k),
                "message": (messages.get(k) or no_detail),
            }
            for k, c in counter.most_common(n)
        ]
        if suite:
            all_items = [i for i in all_items if suite.lower() in (i["suite"] or "").lower()]
        if name:
            all_items = [i for i in all_items if name.lower() in (i["test_name"] or "").lower()]
    else:
        tests = _filter_tests_by_source(tests_items, source)
        all_items = _aggregate_top_failing_tests(
            tests,
            top_n=n,
            suite_sub=suite,
            name_sub=name,
            message_max=300,
        )
        # Attach source for UI badges (single-source modes).
        for it in all_items:
            it.setdefault("source", src if src not in ("real", "synthetic") else None)

    total = len(all_items)
    start = (page - 1) * per_page
    end = start + per_page
    return {
        "items": all_items[start:end],
        "page": page,
        "per_page": per_page,
        "total": total,
        "has_more": end < total,
    }


async def api_services(page: int = 1, per_page: int = 50, status: str = ""):
    snap = await _load_snapshot_async()
    if snap is None:
        raise HTTPException(404, "No snapshot data found.")

    items = snap.services
    if status:
        raw = (status or "").strip().lower()
        if raw == "problems":
            items = [s for s in items if s.status_normalized in ("down", "degraded")]
        else:
            want = normalize_service_status(status)
            items = [s for s in items if s.status_normalized == want]

    total = len(items)
    start = (page - 1) * per_page
    end = start + per_page
    return {
        "items": [json.loads(s.model_dump_json()) for s in items[start:end]],
        "page": page,
        "per_page": per_page,
        "total": total,
        "has_more": end < total,
    }


@app.get("/api/trends", response_class=JSONResponse)
async def api_trends(days: int = 14):
    """Return daily summary history for trend charts (short TTL in-memory cache)."""
    cache_key = f"trends:{days}:{_data_revision}"
    cached = _mem_cache_get(cache_key)
    if cached is not None:
        return JSONResponse(
            content=cached,
            headers={
                "ETag": f'W/"tr-{_data_revision}-{days}"',
                "Cache-Control": "private, max-age=15",
            },
        )
    data = _trends_compute(days)
    _mem_cache_set(cache_key, data)
    return JSONResponse(
        content=data,
        headers={
            "ETag": f'W/"tr-{_data_revision}-{days}"',
            "Cache-Control": "private, max-age=15",
        },
    )


@app.get("/api/uptime", response_class=JSONResponse)
async def api_uptime(days: int = 30):
    """Per-service uptime for last N days (cached)."""
    cache_key = f"uptime:{days}:{_data_revision}"
    cached = _mem_cache_get(cache_key)
    if cached is not None:
        return JSONResponse(
            content=cached,
            headers={
                "ETag": f'W/"up-{_data_revision}-{days}"',
                "Cache-Control": "private, max-age=15",
            },
        )
    data = _uptime_compute(days)
    _mem_cache_set(cache_key, data)
    return JSONResponse(
        content=data,
        headers={
            "ETag": f'W/"up-{_data_revision}-{days}"',
            "Cache-Control": "private, max-age=15",
        },
    )


@app.get("/api/db/stats", response_class=JSONResponse)
async def api_db_stats():
    """SQLite database diagnostics."""
    if not _SQLITE_AVAILABLE:
        return {"enabled": False, "reason": "db.py module not loaded"}
    return db_stats()


@app.get("/api/sources", response_class=JSONResponse)
async def api_sources():
    """Return distinct CI sources present in the snapshot (for filter dropdowns)."""
    snap = _load_snapshot()
    if snap is None:
        return []
    cfg = _load_yaml_config()

    enabled_jenkins = any(
        inst.get("enabled", True) and str(inst.get("url", "") or "").strip()
        for inst in (cfg.get("jenkins_instances", []) or [])
    )
    enabled_gitlab = any(
        inst.get("enabled", True) and str(inst.get("url", "") or "").strip()
        for inst in (cfg.get("gitlab_instances", []) or [])
    )

    sources = {b.source for b in snap.builds if _is_snapshot_build_enabled(b, cfg)}
    if "jenkins" in sources and not enabled_jenkins:
        sources.discard("jenkins")
    if "gitlab" in sources and not enabled_gitlab:
        sources.discard("gitlab")
    return sorted(sources)


@app.get("/api/notifications", response_class=JSONResponse)
async def api_notifications(since_id: int = 0, limit: int = 50):
    """Return state-change notifications (OK→FAIL / FAIL→OK).
    since_id=0 returns all; pass the last seen id to get only new ones.
    """
    items = [n for n in _notifications if n["id"] > since_id]
    return {"items": items[-limit:], "total": len(items), "max_id": _notify_id_seq}


@app.get("/api/events/persisted", response_class=JSONResponse)
async def api_events_persisted(limit: int = 250):
    """State-change timeline persisted across restarts (from data/event_feed.json)."""
    lim = max(1, min(limit, 500))
    return {"items": _event_feed_load(lim)}


@app.get("/api/analytics/sparklines", response_class=JSONResponse)
async def api_analytics_sparklines(jobs: str = "", limit_per_job: int = 12):
    """Per-job duration history from SQLite for dashboard sparklines (batch, max 40 jobs)."""
    if not _SQLITE_AVAILABLE or _db_build_duration is None:
        return {}
    n = max(3, min(limit_per_job, 30))
    names = [j.strip() for j in jobs.split(",") if j.strip()][:40]
    out: dict[str, list[dict]] = {}
    for name in names:
        pts = _db_build_duration(name, n)
        if pts:
            out[name] = pts
    return out


@app.get("/api/analytics/flaky", response_class=JSONResponse)
async def api_analytics_flaky(
    threshold: float = 0.4, min_runs: int = 4, days: int = 30
):
    """Flaky jobs from SQLite history (complements client-side snapshot heuristic)."""
    if not _SQLITE_AVAILABLE or _db_flaky_analysis is None:
        return {"items": [], "source": "none"}
    items = _db_flaky_analysis(threshold, min_runs, days)
    return {"items": items, "source": "sqlite"}


# ── CSV / XLSX Export ─────────────────────────────────────────────────────

def _to_csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")   # BOM for Excel


def _to_xlsx_bytes(headers: list[str], rows: list[list], sheet_name: str = "Data") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(501, "openpyxl not installed — install it with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            # Excel does not support timezone-aware datetimes — strip tzinfo
            if hasattr(val, "tzinfo") and val.tzinfo is not None:
                val = val.replace(tzinfo=None)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width columns (capped at 60)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_builds(
    fmt: str = "csv",
    source: str = "",
    status: str = "",
    job: str = "",
    hours: int = 0,
):
    """Export build records as CSV or XLSX. ?fmt=csv|xlsx"""
    snap = _load_snapshot()
    if not snap:
        raise HTTPException(404, "No snapshot data")

    items = snap.builds
    if source:
        items = [b for b in items if b.source.lower() == source.lower()]
    if status:
        want = normalize_build_status(status)
        items = [b for b in items if b.status_normalized == want]
    if job:
        items = [b for b in items if job.lower() in b.job_name.lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [b for b in items if b.started_at and
                 b.started_at.replace(tzinfo=timezone.utc if b.started_at.tzinfo is None else b.started_at.tzinfo) >= cutoff]

    headers = ["source", "job_name", "build_number", "status", "branch", "started_at", "duration_seconds", "critical", "url"]
    rows = [[getattr(b, h, "") or "" for h in headers] for b in items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = _to_xlsx_bytes(headers, rows, "Builds")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=builds_{date_str}.xlsx"},
        )
    data = _to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=builds_{date_str}.csv"},
    )


async def export_tests(
    fmt: str = "csv",
    status: str = "",
    suite: str = "",
    name: str = "",
    hours: int = 0,
    source: str = "",
):
    """Export test records as CSV or XLSX. ?fmt=csv|xlsx"""
    snap = _load_snapshot()
    if not snap:
        raise HTTPException(404, "No snapshot data")

    items = snap.tests
    if status:
        want = normalize_test_status(status)
        items = [t for t in items if t.status_normalized == want]
    if suite:
        items = [t for t in items if suite.lower() in (t.suite or "").lower()]
    if name:
        items = [t for t in items if name.lower() in t.test_name.lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [t for t in items if t.timestamp and
                 t.timestamp.replace(tzinfo=timezone.utc if t.timestamp.tzinfo is None else t.timestamp.tzinfo) >= cutoff]
    if source:
        items = _filter_tests_by_source(items, source)

    headers = ["source", "suite", "test_name", "status", "duration_seconds", "timestamp", "failure_message", "file_path"]
    rows = [[getattr(t, h, "") or "" for h in headers] for t in items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = _to_xlsx_bytes(headers, rows, "Tests")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=tests_{date_str}.xlsx"},
        )
    data = _to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=tests_{date_str}.csv"},
    )


async def export_failures(
    fmt: str = "csv",
    n: int = 500,
    suite: str = "",
    name: str = "",
    source: str = "",
    hours: int = 0,
    days: int = 0,
):
    """Export top-failures aggregation as CSV or XLSX. ?fmt=csv|xlsx"""
    snap = _load_snapshot()
    if not snap:
        raise HTTPException(404, "No snapshot data")

    tests_win = _filter_tests_by_lookback_hours(snap.tests, hours=int(hours or 0), days=int(days or 0))
    agg = _aggregate_top_failing_tests(
        _filter_tests_by_source(tests_win, source),
        top_n=n,
        suite_sub=suite,
        name_sub=name,
        message_max=500,
    )
    all_items = [(r["test_name"], r["count"], r.get("suite") or "", r.get("message") or "") for r in agg]

    headers = ["test_name", "failure_count", "suite", "last_error"]
    rows = [[i[0], i[1], i[2], i[3]] for i in all_items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = _to_xlsx_bytes(headers, rows, "Failures")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=failures_{date_str}.xlsx"},
        )
    data = _to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=failures_{date_str}.csv"},
    )


async def collect_status():
    """Return background collection state."""
    # When auto-collect was enabled but we haven't collected yet, estimate ETA from enable time.
    next_in = None
    if _auto_collect_enabled and not _collect_state.get("is_collecting"):
        interval = int(_collect_state.get("interval_seconds") or 300)
        if _collect_state.get("last_collected_at"):
            next_in = max(0, interval - int(
                (datetime.now(tz=timezone.utc) -
                 datetime.fromisoformat(_collect_state["last_collected_at"]))
                .total_seconds()
            ))
        elif _auto_collect_enabled_at_iso:
            try:
                enabled_at = datetime.fromisoformat(str(_auto_collect_enabled_at_iso))
                if enabled_at.tzinfo is None:
                    enabled_at = enabled_at.replace(tzinfo=timezone.utc)
                else:
                    enabled_at = enabled_at.astimezone(timezone.utc)
                next_in = max(0, interval - int((datetime.now(tz=timezone.utc) - enabled_at).total_seconds()))
            except Exception:
                next_in = None
    return {
        **_collect_state,
        "auto_collect_enabled": bool(_auto_collect_enabled),
        "next_collect_in_seconds": next_in,
    }


async def set_auto_collect(request: Request):
    """
    Enable/disable the background auto-collect loop (server-wide).
    Body: {"enabled": true|false}
    """
    global _auto_collect_enabled
    rid = _rid(request)
    try:
        body = await request.json()
    except Exception:
        body = {}
    enabled = bool(isinstance(body, dict) and body.get("enabled") in (True, "true", "1", 1))
    _auto_collect_enabled = enabled
    global _auto_collect_enabled_at_iso
    _auto_collect_enabled_at_iso = datetime.now(tz=timezone.utc).isoformat() if enabled else None
    logger.info("[%s] auto-collect set to %s", rid, "on" if enabled else "off")
    # When enabling, start the first collection immediately so UI has data and ETA.
    if enabled:
        try:
            cfg = _load_yaml_config()
        except Exception:
            cfg = None
        if cfg and not _collect_state.get("is_collecting"):
            asyncio.create_task(_do_collect(cfg, force_full=False))
    return {"ok": True, "enabled": bool(_auto_collect_enabled)}


async def collect_logs(limit: int = 400, offset: int = 0):
    """
    Recent collect progress lines for UI live log view.
    Returns {items:[...], total:int}
    """
    try:
        lim = max(1, min(2000, int(limit)))
    except Exception:
        lim = 400
    try:
        off = max(0, int(offset))
    except Exception:
        off = 0
    items = list(_collect_logs)
    total = len(items)
    if off:
        items = items[off:]
    items = items[-lim:]
    return {"items": items, "total": total}


async def collect_slow(limit: int = 10):
    """Top slow (job/build) operations observed during current collect."""
    try:
        lim = max(1, min(100, int(limit)))
    except Exception:
        lim = 10
    items = list(_collect_slow)
    items.sort(key=lambda x: int(x.get("elapsed_ms") or 0), reverse=True)
    return {"items": items[:lim]}


async def trigger_collect(request: Request):
    """Manually trigger a data collection."""
    rid = _rid(request)
    if _collect_state["is_collecting"]:
        logger.info("[%s] collect rejected: already in progress", rid)
        return {"ok": False, "message": "Collection already in progress."}
    force_full = False
    try:
        body = await request.json()
        if isinstance(body, dict) and body.get("force_full") in (True, "true", "1", 1):
            force_full = True
    except Exception:
        force_full = False
    cfg = _load_yaml_config()
    logger.info("[%s] manual collect started", rid)
    asyncio.create_task(_do_collect(cfg, force_full=force_full))
    return {"ok": True, "message": "Collection started."}


async def api_get_settings():
    """Return config for the settings UI — secrets are masked; use POST with same shape to save."""
    return _mask_settings_for_response(_load_yaml_config())


async def api_get_settings_public():
    """Minimal non-secret fields (safe for embedding / diagnostics)."""
    return _public_settings_payload(_load_yaml_config())


async def api_save_settings(request: Request):
    """Persist new settings to config.yaml and restart the collect loop."""
    global _collect_task
    try:
        new_cfg = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON payload")

    saved = _load_yaml_config()
    merged = _merge_settings_secrets(new_cfg, saved)

    p = _config_yaml_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as fh:
        yaml.dump(merged, fh, allow_unicode=True, default_flow_style=False, sort_keys=False)
    logger.info("Settings saved to %s", p.resolve())

    # Cancel the running collect loop
    if _collect_task and not _collect_task.done():
        _collect_task.cancel()
        try:
            await _collect_task
        except asyncio.CancelledError:
            pass
        _collect_task = None

    _collect_state["is_collecting"] = False
    _collect_state["last_error"] = None

    # Restart collect loop with updated config
    w_cfg = merged.get("web", {})
    interval = int(w_cfg.get("collect_interval_seconds", 300))
    _collect_state["interval_seconds"] = interval
    if w_cfg.get("auto_collect", True):
        _collect_task = asyncio.create_task(_collect_loop(merged))
    else:
        asyncio.create_task(_do_collect(merged, force_full=False))

    msg = "Settings saved. Collection restarted with the new configuration."
    cursor_proxy: dict = {}
    try:
        cursor_proxy = await asyncio.to_thread(sync_cursor_proxy_from_config, merged)
        if cursor_proxy.get("message"):
            msg += " " + cursor_proxy["message"]
    except Exception as exc:
        logger.warning("sync_cursor_proxy_from_config after save: %s", exc)
        cursor_proxy = {"managed": False, "ok": False, "message": str(exc)}

    return {"ok": True, "message": msg, "cursor_proxy": cursor_proxy}


def _ui_lang_from_config() -> str:
    cfg = _load_yaml_config()
    gen = MonitorGeneralConfig.model_validate(cfg.get("general") or {})
    lang = str(gen.ui_language).strip().lower()[:5]
    return lang if lang in ("ru", "en") else "en"


async def settings_page(request: Request):
    resp = templates.TemplateResponse(
        "settings.html",
        {"request": request, "ui_language": _ui_lang_from_config()},
    )
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "base-uri 'self'; "
        "object-src 'none'; "
        "frame-ancestors 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://esm.sh; "
        "connect-src 'self'; "
        "font-src 'self' data:; "
    )
    return resp


# ── Action endpoints (trigger builds / restart containers) ────────────────

def _find_jenkins_instance(cfg: dict, url_hint: str) -> dict | None:
    """Find a Jenkins instance from config that matches the given URL prefix."""
    for inst in cfg.get("jenkins_instances", []):
        if inst.get("url", "").rstrip("/") == url_hint.rstrip("/"):
            return inst
    insts = cfg.get("jenkins_instances", [])
    return insts[0] if insts else None


def _find_gitlab_instance(cfg: dict, url_hint: str) -> dict | None:
    for inst in cfg.get("gitlab_instances", []):
        if inst.get("url", "").rstrip("/") == url_hint.rstrip("/"):
            return inst
    insts = cfg.get("gitlab_instances", [])
    return insts[0] if insts else None


@app.post("/api/action/jenkins/build", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def action_jenkins_build(request: Request):
    """Trigger a Jenkins job build.
    Body: {"job_name": "...", "instance_url": "..."}
    """
    rid = _rid(request)
    body = await request.json()
    job_name = body.get("job_name", "")
    instance_url = body.get("instance_url", "")
    if not job_name:
        raise HTTPException(400, "job_name is required")
    _check_rate_limit(f"jenkins:{job_name}")
    logger.info("[%s] action jenkins build job=%s", rid, job_name)

    cfg = _load_yaml_config()
    inst = _find_jenkins_instance(cfg, instance_url)
    if not inst:
        raise HTTPException(404, "No Jenkins instance found in config")

    from clients.jenkins_client import JenkinsClient
    try:
        client = JenkinsClient(
            url=inst["url"],
            username=inst.get("username", ""),
            token=inst.get("token", ""),
            verify_ssl=bool(inst.get("verify_ssl", True)),
        )
        result = client.trigger_build(job_name)
        return result
    except Exception as exc:
        logger.error("Jenkins trigger failed: %s", exc)
        raise HTTPException(500, f"Failed to trigger build: {exc}")


@app.post("/api/action/gitlab/pipeline", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def action_gitlab_pipeline(request: Request):
    """Trigger a GitLab pipeline.
    Body: {"project_id": "...", "ref": "main", "instance_url": "..."}
    """
    rid = _rid(request)
    body = await request.json()
    project_id = body.get("project_id", "")
    ref = body.get("ref", "main")
    instance_url = body.get("instance_url", "")
    if not project_id:
        raise HTTPException(400, "project_id is required")
    _check_rate_limit(f"gitlab:{project_id}:{ref}")
    logger.info("[%s] action gitlab pipeline project=%s ref=%s", rid, project_id, ref)

    cfg = _load_yaml_config()
    inst = _find_gitlab_instance(cfg, instance_url)
    if not inst:
        raise HTTPException(404, "No GitLab instance found in config")

    from clients.gitlab_client import GitLabClient
    try:
        client = GitLabClient(
            url=inst.get("url", "https://gitlab.com"),
            token=inst.get("token", ""),
            verify_ssl=bool(inst.get("verify_ssl", True)),
        )
        result = client.trigger_pipeline(project_id, ref=ref)
        return result
    except Exception as exc:
        logger.error("GitLab trigger failed: %s", exc)
        raise HTTPException(500, f"Failed to trigger pipeline: {exc}")


@app.post("/api/action/docker/container", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def action_docker_container(request: Request):
    """Start, stop, or restart a Docker container.
    Body: {"container_name": "...", "action": "start"|"stop"|"restart"}
    """
    rid = _rid(request)
    body = await request.json()
    container_name = body.get("container_name", "")
    action = (body.get("action") or "restart").lower().strip()
    if not container_name:
        raise HTTPException(400, "container_name is required")
    if action not in ("start", "stop", "restart"):
        raise HTTPException(400, "action must be one of: start, stop, restart")
    _check_rate_limit(f"docker:{container_name}:{action}", window=5)
    logger.info("[%s] action docker %s %s", rid, action, container_name)

    from docker_monitor.monitor import DockerMonitor
    try:
        return DockerMonitor.container_action(container_name, action)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        logger.error("Docker %s failed: %s", action, exc)
        raise HTTPException(500, f"Failed to {action} container: {exc}") from exc


@app.post("/api/action/docker/restart", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def action_docker_restart(request: Request):
    """Backward-compatible alias: restart only."""
    body = await request.json()
    container_name = body.get("container_name", "")
    if not container_name:
        raise HTTPException(400, "container_name is required")
    from docker_monitor.monitor import DockerMonitor
    try:
        return DockerMonitor.container_action(container_name, "restart")
    except Exception as exc:
        logger.error("Docker restart failed: %s", exc)
        raise HTTPException(500, f"Failed to restart container: {exc}") from exc


# ── Log viewers (Jenkins / GitLab / Docker) ───────────────────────────────

@app.get("/api/logs/jenkins", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def api_logs_jenkins(
    request: Request, job_name: str, build_number: int, instance_url: str = ""
):
    """Fetch Jenkins console text for a build. Tries all configured instances unless instance_url is set."""
    if not job_name.strip() or build_number < 1:
        raise HTTPException(400, "job_name and build_number are required")
    _check_rate_limit(f"log:jenkins:{job_name}:{build_number}", window=2)
    logger.info("[%s] GET jenkins log %s #%s", _rid(request), job_name, build_number)

    cfg = _load_yaml_config()
    if instance_url:
        inst = _find_jenkins_instance(cfg, instance_url)
        insts = [inst] if inst else []
    else:
        insts = [i for i in cfg.get("jenkins_instances", []) if i.get("enabled", True)]

    last_err: str | None = None
    from clients.jenkins_client import JenkinsClient
    for inst in insts:
        try:
            client = JenkinsClient(
                url=inst["url"],
                username=inst.get("username", ""),
                token=inst.get("token", ""),
                verify_ssl=bool(inst.get("verify_ssl", True)),
            )
            text = client.fetch_console_text(job_name, build_number)
            return {"ok": True, "log": text, "instance": inst.get("url", "")}
        except Exception as exc:
            last_err = str(exc)
            logger.debug("Jenkins log for %s #%s on %s: %s", job_name, build_number, inst.get("url"), exc)

    raise HTTPException(502, detail=last_err or "Could not fetch Jenkins console log")


@app.get("/api/logs/gitlab", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def api_logs_gitlab(
    request: Request, project_id: str, pipeline_id: int, instance_url: str = ""
):
    """Fetch concatenated GitLab job traces for a pipeline."""
    if not project_id.strip() or pipeline_id < 1:
        raise HTTPException(400, "project_id and pipeline_id are required")
    _check_rate_limit(f"log:gitlab:{project_id}:{pipeline_id}", window=2)
    logger.info("[%s] GET gitlab log %s pipeline %s", _rid(request), project_id, pipeline_id)

    cfg = _load_yaml_config()
    if instance_url:
        inst = _find_gitlab_instance(cfg, instance_url)
        insts = [inst] if inst else []
    else:
        insts = [i for i in cfg.get("gitlab_instances", []) if i.get("enabled", True)]

    last_err: str | None = None
    from clients.gitlab_client import GitLabClient
    for inst in insts:
        try:
            client = GitLabClient(
                url=inst.get("url", "https://gitlab.com"),
                token=inst.get("token", ""),
            )
            text = client.fetch_pipeline_logs(project_id, pipeline_id)
            return {"ok": True, "log": text, "instance": inst.get("url", "")}
        except Exception as exc:
            last_err = str(exc)
            logger.debug("GitLab log pipeline %s on %s: %s", pipeline_id, inst.get("url"), exc)

    raise HTTPException(502, detail=last_err or "Could not fetch GitLab pipeline logs")


@app.get("/api/logs/diff", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def api_logs_diff(
    source: str,
    job_name: str,
    build_number: int,
    instance_url: str = "",
):
    """Compare log of the given build against the last successful build of the same job.
    Returns: {ok, current_build, prev_build, diff_lines: [{tag, line}]}
    tag: '+' added, '-' removed, ' ' unchanged, '?' context only shown
    """
    _check_rate_limit(f"diff:{source}:{job_name}:{build_number}", window=5)
    import difflib

    # Reference build from snapshot: prefer last successful, else last any other build.
    # If snapshot doesn't contain *any* other build for this job, we may fall back to the CI API
    # (at least for Jenkins) because snapshot typically stores only the latest build per job.
    snap = _load_snapshot()
    if snap is None:
        raise HTTPException(404, "No snapshot data")

    def _same_job_rows() -> list:
        return [
            b
            for b in snap.builds
            if b.source.lower() == source.lower()
            and b.job_name == job_name
            and b.build_number != build_number
        ]

    prev_build_number: int | None = None
    reference_kind = ""
    reference_status: str | None = None

    same_job_success = [b for b in _same_job_rows() if b.status_normalized == "success"]
    same_job_any = _same_job_rows()
    if same_job_success:
        same_job_success.sort(key=lambda b: b.started_at or datetime.min, reverse=True)
        prev_build_number = int(same_job_success[0].build_number)
        reference_kind = "last_success"
        reference_status = same_job_success[0].status
    elif same_job_any:
        same_job_any.sort(key=lambda b: b.started_at or datetime.min, reverse=True)
        prev_build_number = int(same_job_any[0].build_number)
        reference_kind = "last_build"
        reference_status = same_job_any[0].status

    cur_text = prev_text = ""
    cfg = _load_yaml_config()
    last_fetch_err: str | None = None

    if source.lower() == "jenkins":
        from clients.jenkins_client import JenkinsClient
        insts = ([_find_jenkins_instance(cfg, instance_url)] if instance_url else
                 [i for i in cfg.get("jenkins_instances", []) if i.get("enabled", True)])
        for inst in (i for i in insts if i):
            try:
                client = JenkinsClient(url=inst.get("url",""), username=inst.get("username",""), token=inst.get("token",""))
                cur_text  = client.fetch_console_text(job_name, build_number)

                # If snapshot couldn't provide a reference build, ask Jenkins directly.
                if prev_build_number is None:
                    ref = client.fetch_reference_build_number(job_name, prefer_success=True)
                    if ref is not None and int(ref) != int(build_number):
                        prev_build_number = int(ref)
                        reference_kind = "jenkins_last_success"
                        reference_status = "success"
                    else:
                        ref2 = client.fetch_reference_build_number(job_name, prefer_success=False)
                        if ref2 is not None and int(ref2) != int(build_number):
                            prev_build_number = int(ref2)
                            reference_kind = "jenkins_last_completed"
                            reference_status = "unknown"

                if prev_build_number is None:
                    raise HTTPException(
                        404,
                        f"No other build for «{job_name}» in snapshot (and no reference build resolved from Jenkins) — run collect to refresh data.",
                    )

                prev_text = client.fetch_console_text(job_name, int(prev_build_number))
                break
            except Exception as exc:
                last_fetch_err = str(exc)
    elif source.lower() == "gitlab":
        from clients.gitlab_client import GitLabClient
        insts = ([_find_gitlab_instance(cfg, instance_url)] if instance_url else
                 [i for i in cfg.get("gitlab_instances", []) if i.get("enabled", True)])
        for inst in (i for i in insts if i):
            try:
                client = GitLabClient(url=inst.get("url",""), token=inst.get("token",""))
                cur_text  = client.fetch_pipeline_logs(job_name, build_number)
                if prev_build_number is None:
                    raise HTTPException(
                        404,
                        f"No other build for «{job_name}» in snapshot — run collect to refresh data.",
                    )
                prev_text = client.fetch_pipeline_logs(job_name, int(prev_build_number))
                break
            except Exception as exc:
                last_fetch_err = str(exc)
    else:
        raise HTTPException(400, f"Diff not supported for source: {source}")

    if not cur_text:
        raise HTTPException(502, "Could not fetch current build log" + (f": {last_fetch_err}" if last_fetch_err else ""))
    if not prev_text:
        raise HTTPException(502, "Could not fetch reference build log" + (f": {last_fetch_err}" if last_fetch_err else ""))

    # Compute unified diff (context=5 lines)
    cur_lines  = cur_text.splitlines()
    prev_lines = prev_text.splitlines()
    diff = list(difflib.unified_diff(prev_lines, cur_lines, lineterm="", n=4))

    return {
        "ok": True,
        "current_build": build_number,
        "reference_build": prev_build_number,
        "reference_status": reference_status,
        "reference_kind": reference_kind,
        "diff": diff,
        "cur_lines": len(cur_lines),
        "prev_lines": len(prev_lines),
    }


@app.get("/api/pipeline/stages", response_class=JSONResponse)
async def api_pipeline_stages(project_id: str, pipeline_id: int, instance_url: str = ""):
    """Return GitLab pipeline job stages with status (lazy-loaded on demand)."""
    if not project_id.strip() or pipeline_id < 1:
        raise HTTPException(400, "project_id and pipeline_id are required")
    _check_rate_limit(f"stages:{project_id}:{pipeline_id}", window=2)

    cfg = _load_yaml_config()
    if instance_url:
        inst = _find_gitlab_instance(cfg, instance_url)
        insts = [inst] if inst else []
    else:
        insts = [i for i in cfg.get("gitlab_instances", []) if i.get("enabled", True)]

    from clients.gitlab_client import GitLabClient
    last_err: str | None = None
    for inst in insts:
        try:
            client = GitLabClient(
                url=inst.get("url", "https://gitlab.com"),
                token=inst.get("token", ""),
            )
            base = inst.get("url", "https://gitlab.com").rstrip("/")
            pid_enc = project_id.replace("/", "%2F")
            resp = client.session.get(
                f"{base}/api/v4/projects/{pid_enc}/pipelines/{pipeline_id}/jobs",
                params={"per_page": 100},
                timeout=client.timeout,
            )
            resp.raise_for_status()
            jobs = resp.json()
            stages: dict[str, list[dict]] = {}
            for j in jobs:
                stage = j.get("stage", "unknown")
                stages.setdefault(stage, []).append({
                    "name":     j.get("name", ""),
                    "status":   j.get("status", "unknown"),
                    "duration": j.get("duration"),
                    "web_url":  j.get("web_url"),
                    "id":       j.get("id"),
                })
            ordered = []
            for stage_name, jobs_list in stages.items():
                ordered.append({"stage": stage_name, "jobs": jobs_list})
            return {"ok": True, "stages": ordered}
        except Exception as exc:
            last_err = str(exc)
    raise HTTPException(502, detail=last_err or "Could not fetch pipeline stages")


@app.get("/api/logs/docker", response_class=JSONResponse, dependencies=[Depends(require_shared_token)])
async def api_logs_docker(container: str, tail: int = 4000):
    """Recent Docker container logs (stdout+stderr)."""
    container = container.strip()
    if not container:
        raise HTTPException(400, "container is required")
    _check_rate_limit(f"log:docker:{container}", window=2)
    from docker_monitor.monitor import DockerMonitor
    try:
        text = DockerMonitor.container_logs_tail(container, tail=max(100, min(tail, 50_000)))
        return {"ok": True, "log": text}
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc
    except Exception as exc:
        logger.error("Docker logs failed: %s", exc)
        raise HTTPException(500, f"Failed to read logs: {exc}") from exc


@app.get("/api/logs/docker/stream", dependencies=[Depends(require_shared_token)])
async def api_logs_docker_stream(container: str):
    """
    Stream Docker logs (like `docker logs -f`). Closes when the client disconnects or the stream ends.
    """
    container = container.strip()
    if not container:
        raise HTTPException(400, "container is required")
    _check_rate_limit(f"log:docker:stream:{container}", window=3)
    from docker_monitor.monitor import DockerMonitor

    def gen():
        yield from DockerMonitor.iter_container_logs_stream(
            container, follow=True, tail=200, timestamps=True
        )

    return StreamingResponse(gen(), media_type="text/plain; charset=utf-8")


@app.post("/webhook/build-complete", dependencies=[Depends(require_shared_token)])
async def webhook_build_complete(request: Request):
    """
    Generic webhook endpoint — receives build completion events from Jenkins/GitLab
    and immediately inserts a record + triggers an incremental collect.

    POST JSON: {
      "source": "jenkins",
      "job": "my-job",
      "status": "success",
      "build_number": 42,
      "url": "https://...",
      "critical": false,
      "trigger_collect": true   // optional: trigger a full collect cycle
    }
    """
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON payload")

    logger.info("Webhook received: %s", payload)

    snap = _load_snapshot() or CISnapshot()
    from models.models import BuildRecord
    record = BuildRecord(
        source=payload.get("source", "webhook"),
        source_instance=(payload.get("source_instance") or None),
        job_name=payload.get("job", "unknown"),
        build_number=payload.get("build_number"),
        status=payload.get("status", "unknown"),
        started_at=datetime.now(tz=timezone.utc),
        url=payload.get("url"),
        critical=payload.get("critical", False),
    )
    snap.builds.insert(0, record)
    save_snapshot(snap)

    # Optionally kick off a full collect cycle (for near-realtime updates)
    if payload.get("trigger_collect", False) and not _collect_state["is_collecting"]:
        cfg = _load_yaml_config()
        asyncio.create_task(_do_collect(cfg, force_full=False))
        return {"ok": True, "message": "Build record added. Full collect triggered."}

    return {"ok": True, "message": "Build record added."}


# ── AI Chat (OpenAI) ──────────────────────────────────────────────────────


def _ai_default_model(provider: str) -> str:
    return {
        "openai": "gpt-4o-mini",
        "gemini": "gemini-2.0-flash",
        "openrouter": "google/gemini-2.0-flash-exp:free",
        "cursor": "auto",
        "ollama": "llama3.1:8b",
        "custom": "gpt-4o-mini",
    }.get(provider, "gpt-4o-mini")


def _looks_like_upstream_unreachable(err_text: str) -> bool:
    low = err_text.lower()
    return any(
        s in low
        for s in (
            "connection refused",
            "failed to connect",
            "errno 111",
            "errno 61",
            "10061",  # Windows: connection refused
            "winerror 10061",
            "name or service not known",
            "getaddrinfo failed",
            "timed out",
            "connect error",
            "connection reset",
        )
    )


def _openai_proxy_url(ai_cfg: dict) -> str | None:
    """Build httpx proxy URL from config (HTTP, HTTPS, SOCKS5). Returns None if disabled."""
    proxy = ai_cfg.get("proxy")
    if not isinstance(proxy, dict) or not proxy.get("enabled"):
        return None
    raw = (proxy.get("url") or "").strip()
    if raw:
        # socks5:// resolves DNS locally; socks5h sends hostname to proxy (needed for api.openai.com on many VPNs)
        low = raw.lower()
        if low.startswith("socks5://") and not low.startswith("socks5h://"):
            raw = "socks5h://" + raw[9:]
        return raw
    host = (proxy.get("host") or "").strip()
    try:
        port = int(proxy.get("port") or 0)
    except (TypeError, ValueError):
        port = 0
    if not host or port <= 0:
        return None
    ptype = (proxy.get("type") or "http").strip().lower()
    if ptype not in ("http", "https", "socks5", "socks5h"):
        ptype = "http"
    # UI "SOCKS5" → socks5h so TLS to api.openai.com uses remote DNS through the tunnel (fixes many geo/VPN cases)
    if ptype == "socks5":
        ptype = "socks5h"
    user = (proxy.get("username") or "").strip()
    password = (proxy.get("password") or "").strip()
    auth = ""
    if user or password:
        auth = f"{quote(user, safe='')}:{quote(password, safe='')}@"
    return f"{ptype}://{auth}{host}:{port}"


async def _http_probe_public_ip(client: httpx.AsyncClient) -> tuple[str | None, str | None]:
    """Return (ip, error_message). Used to verify proxy egress vs direct connection."""
    last_err = "unknown"
    try:
        r = await client.get("https://api.ipify.org", params={"format": "json"}, timeout=20.0)
        r.raise_for_status()
        ip = r.json().get("ip")
        if ip:
            return str(ip), None
        last_err = "ipify returned no ip"
    except Exception as exc:
        last_err = str(exc)
    try:
        r = await client.get("https://icanhazip.com", timeout=20.0)
        r.raise_for_status()
        line = (r.text or "").strip().splitlines()[0].strip()
        if line:
            return line, None
        last_err = "icanhazip empty body"
    except Exception as exc:
        last_err = str(exc)
    return None, last_err


async def api_chat(request: Request):
    """Stream an AI-assistant response powered by OpenAI (or compatible API)."""
    cfg = _load_yaml_config()
    ai_cfg = cfg.get("openai", {})
    provider = ai_cfg.get("provider", "openai")
    api_key = ai_cfg.get("api_key", "").strip()
    if not api_key:
        if provider == "cursor":
            # cursor-api-proxy accepts any key unless CURSOR_BRIDGE_API_KEY is set
            api_key = "unused"
        elif provider == "ollama":
            # Local Ollama OpenAI-compatible API usually needs no key
            api_key = "ollama"
        else:
            raise HTTPException(
                400,
                "API key not configured. Go to Settings → AI Assistant and enter your key.",
            )

    body = await request.json()
    user_messages = body.get("messages", [])
    if not user_messages:
        raise HTTPException(400, "messages list is required")

    context_text = body.get("context", "")
    model = (ai_cfg.get("model") or "").strip() or _ai_default_model(provider)

    _PROVIDER_BASES = {
        "gemini": "https://generativelanguage.googleapis.com/v1beta/openai/",
        "openrouter": "https://openrouter.ai/api/v1",
        # OpenAI-compatible surface from cursor-api-proxy (Cursor Agent CLI + local server)
        "cursor": "http://127.0.0.1:8765/v1",
        "ollama": "http://127.0.0.1:11434/v1",
    }

    system_prompt = (
        "You are an AI assistant embedded in a CI/CD monitoring dashboard. "
        "You help engineers understand build failures, test results, service statuses, "
        "Docker container issues, and CI/CD logs.\n"
        "Rules:\n"
        "- Be concise and actionable.\n"
        "- When analyzing logs, highlight errors, root causes, and suggest fixes.\n"
        "- Use markdown formatting (bold, code blocks, lists) for readability.\n"
        "- Answer in the same language the user writes in.\n"
        "- When a Docker container is down or misbehaving, mention its name clearly so the dashboard can offer a restart button.\n"
        "- When a CI job needs re-running, mention the job name clearly so the dashboard can offer a re-run button.\n"
        "- If the user asks to collect/refresh data, say 'collect' or 'refresh data'.\n"
    )
    if context_text:
        system_prompt += "\nCurrent dashboard context:\n" + context_text[:12000]

    messages = [{"role": "system", "content": system_prompt}]
    for m in user_messages[-20:]:
        role = m.get("role", "user")
        if role in ("user", "assistant"):
            messages.append({"role": role, "content": m.get("content", "")})

    from openai import AsyncOpenAI

    proxy_url = _openai_proxy_url(ai_cfg)
    base_url = (ai_cfg.get("base_url") or "").strip()
    if not base_url:
        base_url = _PROVIDER_BASES.get(provider, "")
    timeout = httpx.Timeout(120.0, connect=60.0)
    px_cfg = ai_cfg.get("proxy") if isinstance(ai_cfg.get("proxy"), dict) else {}
    if px_cfg.get("enabled") and not proxy_url:
        logger.warning(
            "OpenAI proxy enabled in config but URL is incomplete — check host/port or full url (config=%s)",
            _config_yaml_path(),
        )

    async def generate():
        http_client: httpx.AsyncClient | None = None
        try:
            if provider == "cursor" and not _resolve_cursor_agent_cached(cfg):
                yield f"data: {json.dumps({'error': CURSOR_AGENT_UNAVAILABLE_MSG})}\n\n"
                yield "data: [DONE]\n\n"
                return
            client_kw: dict = {"api_key": api_key}
            if base_url:
                client_kw["base_url"] = base_url
            if proxy_url:
                # trust_env=False: avoid HTTP_PROXY / ALL_PROXY overriding or mixing with explicit proxy
                http_client = httpx.AsyncClient(
                    proxy=proxy_url, timeout=timeout, trust_env=False
                )
                client_kw["http_client"] = http_client
                logger.info(
                    "OpenAI chat using explicit proxy (scheme=%s)",
                    proxy_url.split("://", 1)[0] if "://" in proxy_url else "?",
                )
            client = AsyncOpenAI(**client_kw)
            stream = await client.chat.completions.create(
                model=model,
                messages=messages,
                stream=True,
                max_tokens=3000,
                temperature=0.4,
            )
            yielded_text = False
            async for chunk in stream:
                delta = chunk.choices[0].delta if chunk.choices else None
                if delta and delta.content:
                    yielded_text = True
                    yield f"data: {json.dumps({'t': delta.content})}\n\n"
            if not yielded_text:
                empty_msg = (
                    "Cursor: пустой ответ от прокси (см. data/cursor_proxy.log). Проверьте CURSOR_API_KEY, "
                    "работу Cursor Agent и лог прокси."
                    if provider == "cursor"
                    else "Модель вернула пустой ответ. Попробуйте ещё раз или смените модель."
                )
                yield f"data: {json.dumps({'error': empty_msg})}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as exc:
            msg = str(exc)
            if provider == "cursor" and _looks_like_upstream_unreachable(msg):
                msg = (
                    "Cursor: не удалось подключиться к LLM (часто это 127.0.0.1:8765 без запущенного прокси). "
                    "У Cursor нет публичного «прямого» HTTP chat API для токена crsr в сторонних приложениях — "
                    "чат в IDE идёт через инфраструктуру Cursor, а ключ из дашборда — для Cloud Agents / CLI. "
                    "Варианты: (1) установить Cursor Agent CLI, выполнить `npx cursor-api-proxy`, "
                    "в окружении процесса прокси задать CURSOR_API_KEY с вашим crsr_ токеном, base URL оставить "
                    "http://127.0.0.1:8765/v1; (2) переключить провайдера на Gemini или OpenRouter — там один ключ в настройках."
                )
            elif "unsupported_country" in msg or "unsupported_country_region_territory" in msg:
                msg += (
                    " — Geo-block: OpenAI still sees a blocked client region. "
                    "Try switching provider to Gemini or OpenRouter (free, no geo-restrictions) in Settings → AI Assistant."
                )
            yield f"data: {json.dumps({'error': msg})}\n\n"
        finally:
            if http_client is not None:
                await http_client.aclose()

    return StreamingResponse(generate(), media_type="text/event-stream")


async def api_chat_status():
    """Check whether AI chat is configured (without exposing the key)."""
    cfg = _load_yaml_config()
    ai_cfg = cfg.get("openai", {})
    prov = ai_cfg.get("provider", "openai")
    has_key = bool(ai_cfg.get("api_key", "").strip()) or prov in ("cursor", "ollama")
    proxy_ok = _openai_proxy_url(ai_cfg) is not None
    px = ai_cfg.get("proxy") if isinstance(ai_cfg.get("proxy"), dict) else {}
    proxy_on = bool(px.get("enabled")) and proxy_ok
    model = (ai_cfg.get("model") or "").strip() or _ai_default_model(prov)
    cursor_agent_path: str | None = None
    if prov == "cursor":
        cursor_agent_path = _resolve_cursor_agent_cached(cfg)
    return {
        "configured": has_key,
        "provider": prov,
        "model": model,
        "proxy_enabled": proxy_on,
        "proxy_misconfigured": bool(px.get("enabled")) and not proxy_ok,
        "config_path": str(_config_yaml_path()),
        "app_build": _APP_BUILD,
        "cursor_proxy_embedded": _cursor_proxy_running(),
        "cursor_proxy_autostart": _cursor_proxy_autostart_enabled(cfg),
        "cursor_agent_found": cursor_agent_path is not None if prov == "cursor" else None,
        "cursor_agent_path": cursor_agent_path if prov == "cursor" else None,
    }


async def api_chat_proxy_check():
    """Compare public IP with vs without the configured OpenAI proxy (debug VPN routing)."""
    _check_rate_limit("chat:proxy-check", window=10.0)

    cfg = _load_yaml_config()
    ai_cfg = cfg.get("openai", {})
    proxy_url = _openai_proxy_url(ai_cfg)
    timeout = httpx.Timeout(30.0, connect=25.0)

    out: dict = {
        "config_path": str(_config_yaml_path()),
        "proxy_active": bool(proxy_url),
        "proxy_scheme": proxy_url.split("://", 1)[0] if proxy_url and "://" in proxy_url else None,
        "direct": None,
        "via_proxy": None,
    }

    try:
        async with httpx.AsyncClient(timeout=timeout, trust_env=True) as dc:
            dip, derr = await _http_probe_public_ip(dc)
            out["direct"] = {"ok": derr is None and bool(dip), "ip": dip, "error": derr}
    except Exception as exc:
        out["direct"] = {"ok": False, "ip": None, "error": str(exc)}

    if proxy_url:
        try:
            async with httpx.AsyncClient(
                proxy=proxy_url, timeout=timeout, trust_env=False
            ) as pc:
                pip, perr = await _http_probe_public_ip(pc)
                out["via_proxy"] = {"ok": perr is None and bool(pip), "ip": pip, "error": perr}
        except Exception as exc:
            out["via_proxy"] = {"ok": False, "ip": None, "error": str(exc)}
    else:
        px = ai_cfg.get("proxy") if isinstance(ai_cfg.get("proxy"), dict) else {}
        if px.get("enabled"):
            out["via_proxy"] = {
                "ok": False,
                "ip": None,
                "error": "Proxy is enabled but host/port (or full URL) is missing or invalid.",
            }

    return out



from web.routes.builds import router as _builds_router
from web.routes.chat import router as _chat_router
from web.routes.collect import router as _collect_router
from web.routes.incident import router as _incident_router
from web.routes.ops import router as _ops_router
from web.routes.services import router as _services_router
from web.routes.settings import router as _settings_router
from web.routes.tests import router as _tests_router

for __r in (
    _ops_router,
    _incident_router,
    _collect_router,
    _builds_router,
    _tests_router,
    _services_router,
    _settings_router,
    _chat_router,
):
    app.include_router(__r)

# ── Dashboard HTML ────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    snap = await _load_snapshot_async()
    cfg = _load_yaml_config()
    lang = _ui_lang_from_config()

    ctx: dict = {
        "request": request,
        "snap": snap,
        "ui_language": lang,
    }
    if snap:
        ctx["builds_ok"] = sum(
            1 for b in snap.builds if b.status_normalized == "success"
        )
        ctx["builds_fail"] = sum(
            1 for b in snap.builds if b.status_normalized == "failure"
        )
        ctx["tests_fail"] = sum(
            1 for t in snap.tests if t.status_normalized in ("failed", "error")
        )
        ctx["svc_down"] = sum(1 for s in snap.services if s.status_normalized == "down")
    resp = templates.TemplateResponse("index.html", ctx)
    # Ensure browsers don't keep an old dashboard JS/HTML while server code changes.
    # (Settings page already uses no-cache headers; apply same to dashboard.)
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "base-uri 'self'; "
        "object-src 'none'; "
        "frame-ancestors 'self'; "
        "img-src 'self' data:; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://esm.sh; "
        "connect-src 'self'; "
        "font-src 'self' data:; "
    )
    return resp
