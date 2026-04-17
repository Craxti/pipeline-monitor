"""CSV/XLSX export helpers for builds/tests/failures."""

from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta, timezone
from typing import Callable, Optional

from fastapi import HTTPException
from fastapi.responses import Response

from models.models import CISnapshot, normalize_build_status, normalize_test_status
from web.services import tests_analytics


def to_csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel


def to_xlsx_bytes(headers: list[str], rows: list[list], sheet_name: str = "Data") -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(501, "openpyxl not installed — install it with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            if hasattr(val, "tzinfo") and getattr(val, "tzinfo", None) is not None:
                val = val.replace(tzinfo=None)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _must_snapshot(load_snapshot: Callable[[], Optional[CISnapshot]]) -> CISnapshot:
    snap = load_snapshot()
    if not snap:
        raise HTTPException(404, "No snapshot data")
    return snap


async def export_builds(
    *,
    load_snapshot: Callable[[], Optional[CISnapshot]],
    fmt: str = "csv",
    source: str = "",
    status: str = "",
    job: str = "",
    hours: int = 0,
) -> Response:
    snap = _must_snapshot(load_snapshot)

    items = snap.builds
    if source:
        items = [b for b in items if (b.source or "").lower() == source.lower()]
    if status:
        want = normalize_build_status(status)
        items = [b for b in items if b.status_normalized == want]
    if job:
        items = [b for b in items if job.lower() in (b.job_name or "").lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [
            b
            for b in items
            if b.started_at
            and b.started_at.replace(
                tzinfo=timezone.utc if b.started_at.tzinfo is None else b.started_at.tzinfo
            )
            >= cutoff
        ]

    headers = [
        "source",
        "job_name",
        "build_number",
        "status",
        "branch",
        "started_at",
        "duration_seconds",
        "critical",
        "url",
    ]
    rows = [[getattr(b, h, "") or "" for h in headers] for b in items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = to_xlsx_bytes(headers, rows, "Builds")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=builds_{date_str}.xlsx"},
        )
    data = to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=builds_{date_str}.csv"},
    )


async def export_tests(
    *,
    load_snapshot: Callable[[], Optional[CISnapshot]],
    fmt: str = "csv",
    status: str = "",
    suite: str = "",
    name: str = "",
    hours: int = 0,
    source: str = "",
) -> Response:
    snap = _must_snapshot(load_snapshot)

    items = snap.tests
    if status:
        want = normalize_test_status(status)
        items = [t for t in items if t.status_normalized == want]
    if suite:
        items = [t for t in items if suite.lower() in (t.suite or "").lower()]
    if name:
        items = [t for t in items if name.lower() in (t.test_name or "").lower()]
    if hours > 0:
        cutoff = datetime.now(tz=timezone.utc) - timedelta(hours=hours)
        items = [
            t
            for t in items
            if t.timestamp
            and t.timestamp.replace(
                tzinfo=timezone.utc if t.timestamp.tzinfo is None else t.timestamp.tzinfo
            )
            >= cutoff
        ]
    if source:
        items = tests_analytics.filter_tests_by_source(items, source)

    headers = [
        "source",
        "suite",
        "test_name",
        "status",
        "duration_seconds",
        "timestamp",
        "failure_message",
        "file_path",
    ]
    rows = [[getattr(t, h, "") or "" for h in headers] for t in items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = to_xlsx_bytes(headers, rows, "Tests")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=tests_{date_str}.xlsx"},
        )
    data = to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=tests_{date_str}.csv"},
    )


async def export_failures(
    *,
    load_snapshot: Callable[[], Optional[CISnapshot]],
    fmt: str = "csv",
    n: int = 500,
    suite: str = "",
    name: str = "",
    source: str = "",
    hours: int = 0,
    days: int = 0,
) -> Response:
    snap = _must_snapshot(load_snapshot)

    tests_win = tests_analytics.filter_tests_by_lookback_hours(
        snap.tests, hours=int(hours or 0), days=int(days or 0)
    )
    agg = tests_analytics.aggregate_top_failing_tests(
        tests_analytics.filter_tests_by_source(tests_win, source),
        top_n=n,
        suite_sub=suite,
        name_sub=name,
        message_max=500,
    )
    all_items = [
        (r["test_name"], r["count"], r.get("suite") or "", r.get("message") or "")
        for r in agg
    ]

    headers = ["test_name", "failure_count", "suite", "last_error"]
    rows = [[i[0], i[1], i[2], i[3]] for i in all_items]
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt.lower() == "xlsx":
        data = to_xlsx_bytes(headers, rows, "Failures")
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=failures_{date_str}.xlsx"},
        )
    data = to_csv_bytes(headers, rows)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=failures_{date_str}.csv"},
    )

"""CSV/XLSX export helpers for builds, tests, and failures.

Binary/CSV response builders still live in ``web.app`` next to their routes.
"""
